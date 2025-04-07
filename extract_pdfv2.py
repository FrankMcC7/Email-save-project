import os
import re
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import glob
from datetime import datetime
import traceback
import warnings
warnings.filterwarnings('ignore')

# Try to import additional libraries
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False
    print("PyMuPDF not available. For better results, install with: pip install PyMuPDF")

try:
    import tabula
    TABULA_AVAILABLE = True
except ImportError:
    TABULA_AVAILABLE = False
    print("Tabula not available. For better table extraction, install with: pip install tabula-py")

try:
    import subprocess
    POPPLER_AVAILABLE = True
except ImportError:
    POPPLER_AVAILABLE = False


def extract_fund_name(pdf_path):
    """
    Extract the fund name from the third line of the first page.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) > 0:
                first_page_text = pdf.pages[0].extract_text()
                if first_page_text:
                    # Split text by newlines and get the third line if available
                    lines = first_page_text.split('\n')
                    
                    # First try getting the third line if it's not empty
                    if len(lines) >= 3 and lines[2].strip():
                        return lines[2].strip()
                    
                    # Try looking for fund name patterns in the first 10 lines
                    for i in range(min(10, len(lines))):
                        line = lines[i].strip()
                        # Skip empty lines or very short lines
                        if len(line) < 3:
                            continue
                            
                        # Check for fund name indicators
                        if any(indicator in line.lower() for indicator in ['fund', 'sicav', 's.c.a', 'l.p.', 'partners group']):
                            return line
        
        # Fallback to extracting text from first few pages and looking for fund name patterns
        all_text = extract_text_from_pdf(pdf_path, max_pages=3)
        
        # Try to find title lines that might contain fund names
        lines = all_text.split('\n')
        for line in lines[:20]:  # Look only in first 20 lines
            line = line.strip()
            if len(line) > 10 and any(indicator in line.lower() for indicator in ['fund', 'sicav', 's.c.a', 'l.p.', 'partners group']):
                return line
        
        # If all else fails, try to match common fund name patterns
        fund_patterns = [
            r"([A-Za-z0-9\s\-\.&]+(?:S\.C\.A\.|SICAV|Fund|L\.P\.))",
            r"([A-Za-z]+\s+[A-Za-z]+\s+[A-Za-z]+(?:\s+[IVX]+)?)"
        ]
        
        for pattern in fund_patterns:
            match = re.search(pattern, all_text)
            if match:
                return match.group(1).strip()
    except Exception as e:
        print(f"Error extracting fund name: {str(e)}")
    
    return "Unknown Fund"


def extract_text_from_pdf(pdf_path, max_pages=None, start_page=0):
    """
    Extract text from PDF, optionally limiting to max_pages starting from start_page.
    """
    try:
        all_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            # Determine which pages to extract
            if max_pages is None:
                pages_to_extract = pdf.pages[start_page:]
            else:
                end_page = min(start_page + max_pages, len(pdf.pages))
                pages_to_extract = pdf.pages[start_page:end_page]
                
            for page in pages_to_extract:
                page_text = page.extract_text() or ""
                all_text += page_text + "\n"
        return all_text
    except Exception as e:
        print(f"Error extracting text from PDF: {str(e)}")
        return ""


def analyze_pdf_structure(pdf_path):
    """
    Analyze PDF to determine its structure and guide extraction approach.
    """
    pdf_info = {
        "has_tables": False,
        "is_image_based": False,
        "has_key_figures_section": False,
        "is_partners_group": False,
        "page_count": 0
    }
    
    try:
        # Try with pdfplumber first
        with pdfplumber.open(pdf_path) as pdf:
            pdf_info["page_count"] = len(pdf.pages)
            
            # Check first 3 pages
            for i in range(min(3, len(pdf.pages))):
                page = pdf.pages[i]
                page_text = page.extract_text() or ""
                
                # Check if this PDF has tables
                if page.find_tables():
                    pdf_info["has_tables"] = True
                
                # Check if this PDF has "Key Figures" section (case-insensitive)
                if "key figures" in page_text.lower():
                    pdf_info["has_key_figures_section"] = True
                
                # Check if this is a Partners Group document
                if "partners group" in page_text.lower():
                    pdf_info["is_partners_group"] = True
                    
                # If page has very little extractable text, might be image-based
                if len(page_text) < 100:
                    pdf_info["is_image_based"] = True
                    
            # If we haven't found tables yet, try a different approach
            if not pdf_info["has_tables"] and PYMUPDF_AVAILABLE:
                doc = fitz.open(pdf_path)
                for i in range(min(3, doc.page_count)):
                    page = doc[i]
                    if page.get_text("dict")["blocks"]:
                        # Check for structures that look like tables
                        blocks = page.get_text("dict")["blocks"]
                        # Look for grid-like text arrangement
                        if any(len(block.get("lines", [])) > 3 for block in blocks):
                            pdf_info["has_tables"] = True
                            
                        # Check for Partners Group and Key Figures mentions
                        page_text = page.get_text()
                        if "partners group" in page_text.lower():
                            pdf_info["is_partners_group"] = True
                        if "key figures" in page_text.lower():
                            pdf_info["has_key_figures_section"] = True
        
        return pdf_info
    except Exception as e:
        print(f"Error analyzing PDF structure: {str(e)}")
        return pdf_info


def extract_text_multi_library(pdf_path):
    """
    Try multiple libraries to extract text from PDF.
    """
    text = ""
    
    # Try pdfplumber first
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                text += page_text + "\n"
        
        if text.strip():
            return text
    except Exception as e:
        print(f"pdfplumber extraction failed: {str(e)}")
    
    # Try PyMuPDF if available
    if PYMUPDF_AVAILABLE:
        try:
            doc = fitz.open(pdf_path)
            text = ""
            for page in doc:
                text += page.get_text() + "\n"
            
            if text.strip():
                return text
        except Exception as e:
            print(f"PyMuPDF extraction failed: {str(e)}")
    
    # If all else fails, use the original extract_text_from_pdf function
    return extract_text_from_pdf(pdf_path)


def extract_tables_multi_library(pdf_path):
    """
    Try multiple libraries to extract tables from PDF.
    """
    all_tables = []
    
    # First check if this is a Partners Group document with Key Figures
    pdf_info = analyze_pdf_structure(pdf_path)
    
    # Try pdfplumber with different settings - prioritize settings based on document type
    table_settings_list = []
    
    # Partners Group documents with Key Figures often have a specific structure
    if pdf_info["is_partners_group"] and pdf_info["has_key_figures_section"]:
        table_settings_list = [
            # Partners Group often has tables defined by text alignment rather than lines
            {'vertical_strategy': 'text', 'horizontal_strategy': 'text', 'snap_tolerance': 5},
            {'vertical_strategy': 'text', 'horizontal_strategy': 'lines'},
            {'vertical_strategy': 'lines', 'horizontal_strategy': 'text'},
            {'vertical_strategy': 'lines', 'horizontal_strategy': 'lines'}
        ]
    else:
        # Default settings for other documents
        table_settings_list = [
            {'vertical_strategy': 'lines', 'horizontal_strategy': 'lines'},
            {'vertical_strategy': 'text', 'horizontal_strategy': 'text'},
            {'vertical_strategy': 'lines', 'horizontal_strategy': 'text'},
            {'vertical_strategy': 'text', 'horizontal_strategy': 'lines'}
        ]
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num in range(min(5, len(pdf.pages))):
                page = pdf.pages[page_num]
                
                for settings in table_settings_list:
                    try:
                        tables = page.extract_tables(settings)
                        if tables:
                            all_tables.extend([{'page': page_num, 'tables': tables, 'source': 'pdfplumber'}])
                            break  # If we found tables with these settings, move to next page
                    except:
                        continue
    except Exception as e:
        print(f"pdfplumber table extraction failed: {str(e)}")
    
    # Try tabula if available
    if TABULA_AVAILABLE and not all_tables:
        try:
            tabula_tables = tabula.read_pdf(pdf_path, pages='1-5', multiple_tables=True)
            if tabula_tables:
                # Convert tabula tables to our format
                for table in tabula_tables:
                    if not table.empty:
                        all_tables.append({'page': 0, 'tables': [table.values.tolist()], 'source': 'tabula'})
        except Exception as e:
            print(f"tabula table extraction failed: {str(e)}")
    
    return all_tables


def clean_number(value_str):
    """
    Clean and convert a string value to a float.
    Handles various number formats including apostrophes, commas, spaces.
    """
    try:
        if value_str is None:
            return None
            
        # If it's already a number, return it
        if isinstance(value_str, (int, float)):
            return float(value_str)
            
        if isinstance(value_str, str):
            # Handle percentage values - skip these
            if "%" in value_str:
                return None
            
            # Skip non-numeric text or extremely short text
            if len(value_str.strip()) == 0 or value_str.strip().isalpha():
                return None
            
            # Check if the string contains any digits at all
            if not any(c.isdigit() for c in value_str):
                return None
            
            # Special handling for Partners Group format with apostrophes (e.g., 595'446'138)
            # First check if it matches the pattern
            if "'" in value_str and re.search(r'\d+(?:\'\d+)+', value_str.strip()):
                # Simply remove all apostrophes
                cleaned = value_str.replace("'", "")
                try:
                    return float(cleaned)
                except:
                    pass
                
            # Standard cleaning for other formats
            cleaned = value_str.replace("'", "").replace(",", "").replace(" ", "").replace("x", "")
            
            # Handle any remaining non-numeric characters (except decimal point)
            final_value = ""
            decimal_found = False
            for char in cleaned:
                if char.isdigit():
                    final_value += char
                elif char == "." and not decimal_found:
                    final_value += char
                    decimal_found = True
                    
            if final_value:
                # Convert to float
                try:
                    value = float(final_value)
                    # Don't filter by size - some NAV values can be small
                    return value
                except:
                    pass
    except:
        pass
        
    return None


def extract_nav_from_tables(tables_data):
    """
    Extract NAV information from tables.
    """
    results = []
    
    for table_entry in tables_data:
        page = table_entry['page']
        tables = table_entry['tables']
        
        for table in tables:
            # Skip empty or tiny tables
            if not table or len(table) < 2:
                continue
            
            # Look for NAV values in this table
            for row_idx, row in enumerate(table):
                if not row or len(row) < 2:
                    continue
                
                row_str = [str(cell).strip() if cell is not None else "" for cell in row]
                row_text = " ".join(row_str).lower()
                
                # Check if this row contains NAV information
                if "net asset value" in row_text or "nav" in row_text.split():
                    # Get values
                    values = []
                    for cell_idx, cell in enumerate(row):
                        if cell_idx > 0:  # Skip first column (label)
                            value = clean_number(str(cell))
                            if value is not None and value > 100:  # NAV values are typically large
                                values.append((cell_idx, value))
                    
                    if len(values) >= 2:
                        # Look for header row (dates)
                        header_row = None
                        for i in range(row_idx):
                            if len(table[i]) >= len(row):
                                header_text = " ".join([str(cell).strip() if cell is not None else "" for cell in table[i]])
                                # Check for date patterns
                                if re.search(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', header_text):
                                    header_row = table[i]
                                    break
                        
                        if header_row:
                            idx1, value1 = values[0]
                            idx2, value2 = values[1]
                            
                            header1 = str(header_row[idx1]).strip() if idx1 < len(header_row) and header_row[idx1] is not None else "Period 1"
                            header2 = str(header_row[idx2]).strip() if idx2 < len(header_row) and header_row[idx2] is not None else "Period 2"
                            
                            results.append({
                                'period1_label': header1,
                                'period1_nav': value1,
                                'period2_label': header2,
                                'period2_nav': value2,
                                'confidence': 0.9,  # High confidence for table with header
                                'source': f'table_p{page}'
                            })
                        else:
                            # No header row found, extract values anyway
                            results.append({
                                'period1_label': "Period 1",
                                'period1_nav': values[0][1],
                                'period2_label': "Period 2",
                                'period2_nav': values[1][1],
                                'confidence': 0.7,  # Medium confidence without header
                                'source': f'table_p{page}'
                            })
    
    # Sort by confidence
    return sorted(results, key=lambda x: x['confidence'], reverse=True)


def extract_nav_with_enhanced_patterns(text):
    """
    Use enhanced patterns to extract NAV values from text.
    Always looking for NAV as row label (never in column headers).
    """
    results = []
    
    # Split text into lines for row-based analysis
    lines = text.split('\n')
    
    # Enhanced NAV row patterns - looking for NAV as a row label followed by values
    nav_row_patterns = [
        # Pattern 1: "Net Asset Value" followed by values
        r"Net\s+[Aa]sset\s+[Vv]alue\s*([\d\s,\.\']+)\s*([\d\s,\.\']+)",
        
        # Pattern 2: "NAV" followed by values
        r"\bNAV\b\s*([\d\s,\.\']+)\s*([\d\s,\.\']+)",
        
        # Pattern 3: More flexible pattern with intervening content
        r"Net\s+[Aa]sset\s+[Vv]alue.*?([\d\s,\.\']+).*?([\d\s,\.\']+)"
    ]
    
    # Date patterns
    date_patterns = [
        r'\b(\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4})\b',
        r'\b(\d{2}\.\d{2}\.\d{4})\b',
        r'\b((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4})\b',
        r'\b((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},?\s+\d{4})\b',
        r'\b(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})\b',
        r'\b(\d{2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4})\b'
    ]
    
    # Extract dates from the text
    dates = []
    for pattern in date_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        dates.extend(matches)
    
    # Ensure we have unique dates
    dates = list(set(dates))
    
    # Sort dates (assuming they are in standard formats)
    dates.sort()
    
    # First try to find NAV rows with values on the same line
    for line in lines:
        if "net asset value" in line.lower() or "nav" in line.lower():
            for pattern in nav_row_patterns:
                matches = re.search(pattern, line, re.IGNORECASE)
                if matches:
                    # Extract the values from the matches
                    if len(matches.groups()) >= 2:
                        value1 = clean_number(matches.group(1))
                        value2 = clean_number(matches.group(2))
                        
                        if value1 is not None and value2 is not None:
                            # If we have dates and values
                            if len(dates) >= 2:
                                results.append({
                                    'period1_label': dates[-2],  # Second most recent date
                                    'period1_nav': value1,
                                    'period2_label': dates[-1],  # Most recent date
                                    'period2_nav': value2,
                                    'confidence': 0.85,  # Good confidence with dates
                                    'source': 'pattern_row_with_dates'
                                })
                                break
                            else:
                                # If we have values but no dates
                                results.append({
                                    'period1_label': "Period 1",
                                    'period1_nav': value1,
                                    'period2_label': "Period 2", 
                                    'period2_nav': value2,
                                    'confidence': 0.75,  # Lower confidence without dates
                                    'source': 'pattern_row_no_dates'
                                })
                                break
    
    # If we haven't found NAV row with values yet, try looking for NAV row and then extract values
    if not results:
        # Find the NAV line index
        nav_line_idx = -1
        for i, line in enumerate(lines):
            if "net asset value" in line.lower() or "nav" in line.lower().split():
                nav_line_idx = i
                break
        
        # If we found a NAV line, look for values in this line and the next few lines
        if nav_line_idx >= 0:
            # Extract all numbers from the NAV line and next two lines
            values = []
            for i in range(nav_line_idx, min(nav_line_idx + 3, len(lines))):
                numbers = re.findall(r"[\d\s,\.\']+", lines[i])
                for num in numbers:
                    value = clean_number(num)
                    if value is not None:
                        values.append(value)
            
            # If we found at least two values
            if len(values) >= 2:
                # If we have dates and values
                if len(dates) >= 2:
                    results.append({
                        'period1_label': dates[-2],  # Second most recent date
                        'period1_nav': values[0],
                        'period2_label': dates[-1],  # Most recent date
                        'period2_nav': values[1],
                        'confidence': 0.7,  # Medium confidence with dates but values across lines
                        'source': 'pattern_multi_line_with_dates'
                    })
                else:
                    # If we have values but no dates
                    results.append({
                        'period1_label': "Period 1",
                        'period1_nav': values[0],
                        'period2_label': "Period 2", 
                        'period2_nav': values[1],
                        'confidence': 0.6,  # Lower confidence without dates and values across lines
                        'source': 'pattern_multi_line_no_dates'
                    })
    
    return results


def direct_table_extraction(pdf_path):
    """
    Directly target the Key Figures table and extract Net asset value.
    Always looking for Net asset value as a row label (never in column headers).
    This approach focuses specifically on the table structure seen in examples.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check pages 1-5 for the key figures table
            for page_num in range(min(5, len(pdf.pages))):
                page = pdf.pages[page_num]
                
                # Get the text to check for Key Figures section
                page_text = page.extract_text() or ""
                
                # Only process pages with "Key figures" or "key figures"
                if "Key figures" in page_text or "key figures" in page_text.lower():
                    # Extract all tables from this page
                    for table_settings in [
                        {'vertical_strategy': 'text', 'horizontal_strategy': 'text'},
                        {'vertical_strategy': 'lines', 'horizontal_strategy': 'lines'},
                        {'vertical_strategy': 'text', 'horizontal_strategy': 'lines'},
                        {'vertical_strategy': 'lines', 'horizontal_strategy': 'text'}
                    ]:
                        try:
                            tables = page.extract_tables(table_settings)
                            
                            # Process each table
                            for table in tables:
                                if not table or len(table) < 3:  # Skip small tables
                                    continue
                                
                                # Find header row with dates
                                header_row = None
                                header_row_idx = -1
                                
                                for i, row in enumerate(table[:5]):
                                    if not row or len(row) < 3:
                                        continue
                                        
                                    row_text = " ".join([str(cell).strip() if cell is not None else "" for cell in row])
                                    # Look for dates or month names in header row
                                    if (re.search(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', row_text) or
                                        "september" in row_text.lower() or "december" in row_text.lower() or 
                                        "march" in row_text.lower() or "june" in row_text.lower()):
                                        header_row = row
                                        header_row_idx = i
                                        break
                                
                                if header_row is None:
                                    continue
                                
                                # Look for NAV row (Net asset value as row label)
                                for row_idx, row in enumerate(table):
                                    if not row or len(row) < 3:
                                        continue
                                        
                                    # Check if first cell contains Net asset value (as row label)
                                    first_cell = str(row[0]).strip().lower() if row[0] is not None else ""
                                    
                                    if "net asset value" in first_cell or "nav" in first_cell.split():
                                        # Found the NAV row - extract values
                                        values = []
                                        for cell_idx in range(1, len(row)):
                                            if row[cell_idx] is not None:
                                                value = clean_number(str(row[cell_idx]))
                                                if value is not None:
                                                    values.append((cell_idx, value))
                                        
                                        # Need at least two values
                                        if len(values) >= 2:
                                            # Get corresponding headers
                                            idx1, value1 = values[0]
                                            idx2, value2 = values[1]
                                            
                                            if header_row:
                                                header1 = str(header_row[idx1]).strip() if idx1 < len(header_row) and header_row[idx1] is not None else "Period 1"
                                                header2 = str(header_row[idx2]).strip() if idx2 < len(header_row) and header_row[idx2] is not None else "Period 2"
                                            else:
                                                # Extract dates from page text if headers not found
                                                date_matches = re.findall(r'\b\d{1,2}[\.\/\s]+(?:[A-Za-z]+|\d{1,2})[\.\/\s]+\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b|\b(?:Q[1-4]|20\d\d)\b', page_text)
                                                if len(date_matches) >= 2:
                                                    header1 = date_matches[0]
                                                    header2 = date_matches[1]
                                                else:
                                                    header1 = "Period 1"
                                                    header2 = "Period 2"
                                            
                                            return header1, value1, header2, value2
                        except Exception as e:
                            # Continue with next table settings if this fails
                            pass
    except Exception as e:
        print(f"Error in direct table extraction: {str(e)}")
        
    return None, None, None, None


def scan_text_for_nav(pdf_path):
    """
    Scan the PDF text for Net Asset Value mentions and extract data.
    This is a fallback approach when table extraction fails.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Extract text from the first few pages
            all_text = ""
            for page_num in range(min(5, len(pdf.pages))):
                page_text = pdf.pages[page_num].extract_text() or ""
                all_text += page_text + "\n"
                
                # Check each page individually first
                lines = page_text.split('\n')
                
                # 1. First, look for "Net asset value" line specifically
                for i, line in enumerate(lines):
                    if "net asset value" in line.lower():
                        # Found NAV line - extract numbers
                        number_matches = re.findall(r"[\d',\.]+", line)
                        values = []
                        
                        for num in number_matches:
                            value = clean_number(num)
                            if value is not None and value > 1000:
                                values.append(value)
                        
                        if len(values) >= 2:
                            # Look for date headers in previous lines
                            date_matches = []
                            for j in range(max(0, i-5), i):
                                date_match = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', lines[j])
                                if date_match:
                                    date_matches.extend(date_match)
                            
                            if len(date_matches) >= 2:
                                return date_matches[0], values[0], date_matches[1], values[1]
                            else:
                                # Look for dates in the entire page
                                date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', page_text)
                                if len(date_matches) >= 2:
                                    return date_matches[0], values[0], date_matches[1], values[1]
                                else:
                                    return "Period 1", values[0], "Period 2", values[1]
            
            # If page-by-page approach failed, try with whole text
            nav_matches = re.finditer(r'(?:[Nn]et\s+[Aa]sset\s+[Vv]alue|NAV)', all_text)
            
            for match in nav_matches:
                # Extract text around the match
                start_pos = max(0, match.start() - 100)
                end_pos = min(len(all_text), match.end() + 300)
                context = all_text[start_pos:end_pos]
                
                # Find dates in the context
                dates = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', context)
                
                # Find large numbers that could be NAV values
                number_matches = re.findall(r"[\d',\.]+", context)
                values = []
                
                for num in number_matches:
                    value = clean_number(num)
                    if value is not None and value > 10000:  # NAV values are typically large
                        values.append(value)
                
                if len(dates) >= 2 and len(values) >= 2:
                    return dates[0], values[0], dates[1], values[1]
    except Exception as e:
        print(f"Error in text scanning: {str(e)}")
        
    return None, None, None, None


def last_resort_extraction(pdf_path):
    """
    Absolute last resort approach - specifically look for patterns like in the Excel screenshot.
    This approach focuses on known formats from the examples provided.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check first 3 pages
            for page_num in range(min(3, len(pdf.pages))):
                page = pdf.pages[page_num]
                page_text = page.extract_text() or ""
                
                # Look specifically for "Net asset value" in a line
                lines = page_text.split('\n')
                
                # First approach: Look for the exact line format with Net asset value
                for line in lines:
                    if "net asset value" in line.lower():
                        # Try to extract exactly two numbers from this line
                        number_matches = re.findall(r"[\d',\.]+", line)
                        values = []
                        
                        for num in number_matches:
                            value = clean_number(num)
                            if value is not None and value > 10000:  # NAV values are typically large
                                values.append(value)
                        
                        if len(values) >= 2:
                            # Extract dates from the page text
                            date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', page_text)
                            
                            if len(date_matches) >= 2:
                                return date_matches[0], values[0], date_matches[1], values[1]
                            else:
                                # If no dates found, just use Period 1 and Period 2
                                return "Period 1", values[0], "Period 2", values[1]
                
                # Second approach: Look for numeric patterns in lines with dates
                date_line_idx = -1
                for i, line in enumerate(lines):
                    if re.search(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', line):
                        date_line_idx = i
                        break
                
                if date_line_idx >= 0:
                    # Found a line with dates, now look for "Net asset value" within the next 15 lines
                    for i in range(date_line_idx + 1, min(date_line_idx + 15, len(lines))):
                        if "net asset value" in lines[i].lower():
                            # Found the NAV line
                            number_matches = re.findall(r"[\d',\.]+", lines[i])
                            values = []
                            
                            for num in number_matches:
                                value = clean_number(num)
                                if value is not None and value > 10000:
                                    values.append(value)
                            
                            if len(values) >= 2:
                                # Extract dates from the date line
                                date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', lines[date_line_idx])
                                
                                if len(date_matches) >= 2:
                                    return date_matches[0], values[0], date_matches[1], values[1]
            
            # Third approach: Process all lines looking for specific formats like in the Excel
            for page_num in range(min(5, len(pdf.pages))):
                page_text = pdf.pages[page_num].extract_text() or ""
                lines = page_text.split('\n')
                
                for line in lines:
                    # Look for lines that match the format "some text number1 number2"
                    if "asset" in line.lower() or "value" in line.lower():
                        number_matches = re.findall(r"[\d',\.]+", line)
                        values = []
                        
                        for num in number_matches:
                            value = clean_number(num)
                            if value is not None and value > 10000:
                                values.append(value)
                        
                        if len(values) >= 2:
                            # Extract dates from the page text
                            date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', page_text)
                            
                            if len(date_matches) >= 2:
                                return date_matches[0], values[0], date_matches[1], values[1]
    except Exception as e:
        print(f"Error in last resort extraction: {str(e)}")
        
    return None, None, None, None


def extract_partners_group_key_figures(pdf_path):
    """
    Special extraction function for Partners Group Key Figures tables.
    This function is specifically designed to handle tables with:
    - Dates as column headers (both formats: "30 September 2024" and "30.09.2024")
    - "Net asset value" as a row label
    - NAV values at the intersection
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check first few pages for "Key figures" section
            for page_num in range(min(5, len(pdf.pages))):
                page = pdf.pages[page_num]
                page_text = page.extract_text() or ""
                
                # Debug for page text
                print(f"\n------ Checking page {page_num+1} ------")
                if "key figures" in page_text.lower():
                    print("Found 'Key figures' on this page!")
                
                # Try different table extraction settings
                for table_settings in [
                    {'vertical_strategy': 'text', 'horizontal_strategy': 'text'},
                    {'vertical_strategy': 'lines', 'horizontal_strategy': 'text'},
                    {'vertical_strategy': 'text', 'horizontal_strategy': 'lines'},
                    {'vertical_strategy': 'lines', 'horizontal_strategy': 'lines'}
                ]:
                    try:
                        tables = page.extract_tables(table_settings)
                        
                        for table_idx, table in enumerate(tables):
                            if not table or len(table) < 3:  # Key figures tables usually have several rows
                                continue
                                
                            print(f"\nExamining Table #{table_idx+1} ({len(table)} rows)")
                            
                            # Dump all table rows to debug
                            for i, row in enumerate(table):
                                if row:
                                    row_str = ' | '.join([str(cell).strip() if cell is not None else "None" for cell in row])
                                    print(f"Row {i}: {row_str}")
                            
                            # First identify the date header row
                            date_header_row = None
                            date_header_idx = -1
                            
                            for i, row in enumerate(table[:8]):  # Check first 8 rows for header (increased from 5)
                                if not row or len(row) < 2:
                                    continue
                                
                                # Convert row to text for checking
                                row_text = " ".join([str(cell).strip() if cell is not None else "" for cell in row])
                                
                                # Check cells individually for date formats
                                date_cells = []
                                for j, cell in enumerate(row):
                                    if cell is not None:
                                        cell_text = str(cell).strip()
                                        
                                        # Check for various date formats with more flexible patterns
                                        if (re.search(r'\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}', cell_text) or
                                            re.search(r'\d{1,2}\s*(?:January|February|March|April|May|June|July|August|September|October|November|December)\s*\d{4}', cell_text, re.IGNORECASE) or
                                            re.search(r'\d{1,2}(?:January|February|March|April|May|June|July|August|September|October|November|December)\d{4}', cell_text, re.IGNORECASE) or
                                            re.search(r'\d{2}\.\d{2}\.\d{4}', cell_text) or
                                            ("september" in cell_text.lower() and re.search(r'\d{4}', cell_text)) or
                                            ("december" in cell_text.lower() and re.search(r'\d{4}', cell_text))):
                                            date_cells.append((j, cell_text))
                                
                                if len(date_cells) >= 2:  # Need at least two date columns
                                    print(f"Found date header at row {i}: {date_cells}")
                                    date_header_row = row
                                    date_header_idx = i
                                    break
                            
                            # If we found date headers, now look for the NAV row
                            if date_header_row is not None:
                                print("Searching for NAV row after date header...")
                                
                                # First find NAV using direct string matches
                                nav_row = None
                                nav_row_idx = -1
                                
                                # Patterns that might indicate a NAV row
                                nav_patterns = [
                                    r"net\s*asset\s*value",
                                    r"\bnav\b",
                                    r"^net",
                                    r"asset.*value"
                                ]
                                
                                # Search through ALL rows after the header
                                for row_idx in range(date_header_idx + 1, len(table)):
                                    row = table[row_idx]
                                    if not row or len(row) < 3:
                                        continue
                                    
                                    # Check the first cell for various NAV indicators
                                    first_cell = str(row[0]).strip().lower() if row[0] is not None else ""
                                    
                                    print(f"  Checking row {row_idx}, first cell: '{first_cell}'")
                                    
                                    # Super flexible matching for NAV indicators
                                    is_nav_row = False
                                    for pattern in nav_patterns:
                                        if re.search(pattern, first_cell, re.IGNORECASE):
                                            is_nav_row = True
                                            break
                                    
                                    if is_nav_row:
                                        print(f"  FOUND NAV ROW at index {row_idx}: {first_cell}")
                                        nav_row = row
                                        nav_row_idx = row_idx
                                        break
                                
                                # If we found the NAV row, extract values
                                if nav_row is not None:
                                    # Extract values from this row
                                    values = []
                                    for i in range(1, len(nav_row)):
                                        if nav_row[i] is not None:
                                            cell_text = str(nav_row[i]).strip()
                                            print(f"  Nav value cell {i}: '{cell_text}'")
                                            
                                            # Try to extract a number
                                            value = clean_number(cell_text)
                                            if value is not None:
                                                print(f"  Extracted value: {value}")
                                                values.append((i, value))
                                    
                                    # If we found at least 2 values
                                    if len(values) >= 2:
                                        idx1, value1 = values[0]
                                        idx2, value2 = values[1]
                                        
                                        # Get corresponding headers
                                        header1 = str(date_header_row[idx1]).strip() if idx1 < len(date_header_row) and date_header_row[idx1] is not None else "Period 1"
                                        header2 = str(date_header_row[idx2]).strip() if idx2 < len(date_header_row) and date_header_row[idx2] is not None else "Period 2"
                                        
                                        print(f"EXTRACTION SUCCESSFUL: {header1}: {value1}, {header2}: {value2}")
                                        return header1, value1, header2, value2
                                else:
                                    print("No explicit NAV row found. Trying fallback methods...")
                                    
                                    # Fallback: Scan all rows for anything that might be NAV
                                    # Look for rows containing numeric values that are different from others
                                    potential_rows = []
                                    
                                    for row_idx in range(date_header_idx + 1, len(table)):
                                        row = table[row_idx]
                                        if not row or len(row) < 3:
                                            continue
                                        
                                        # Skip rows with percentages (%) or "commitments"
                                        first_cell = str(row[0]).strip().lower() if row[0] is not None else ""
                                        if "%" in first_cell or "commitment" in first_cell or "capital" in first_cell:
                                            continue
                                        
                                        # Extract values from this row
                                        row_values = []
                                        for i in range(1, len(row)):
                                            if row[i] is not None:
                                                value = clean_number(str(row[i]))
                                                if value is not None:
                                                    row_values.append((i, value))
                                        
                                        # If this row has at least 2 values, consider it
                                        if len(row_values) >= 2:
                                            potential_rows.append((row_idx, row, row_values))
                                    
                                    # If we found potential NAV rows, use the one with the largest values
                                    # (NAV is typically larger than other metrics)
                                    if potential_rows:
                                        # Sort by average value in descending order
                                        potential_rows.sort(key=lambda x: sum(val for _, val in x[2])/len(x[2]), reverse=True)
                                        
                                        # Use the row with largest average values
                                        _, best_row, best_values = potential_rows[0]
                                        
                                        print(f"Using potential NAV row with largest values: {best_row[0]}")
                                        
                                        idx1, value1 = best_values[0]
                                        idx2, value2 = best_values[1]
                                        
                                        # Get corresponding headers
                                        header1 = str(date_header_row[idx1]).strip() if idx1 < len(date_header_row) and date_header_row[idx1] is not None else "Period 1"
                                        header2 = str(date_header_row[idx2]).strip() if idx2 < len(date_header_row) and date_header_row[idx2] is not None else "Period 2"
                                        
                                        print(f"FALLBACK EXTRACTION: {header1}: {value1}, {header2}: {value2}")
                                        return header1, value1, header2, value2
                    except Exception as e:
                        print(f"Error processing table: {str(e)}")
                        continue
    except Exception as e:
        print(f"Error in Partners Group key figures extraction: {str(e)}")
    
    print("All extraction attempts failed for this file")
    return None, None, None, None


def scan_for_nav_row(page_text, fund_name):
    """
    Fallback method to extract NAV values directly from text.
    This is a last resort when table extraction fails.
    """
    try:
        print("\nTrying direct text scanning for NAV rows...")
        lines = page_text.split('\n')
        
        # First look for date headers in the text
        date_lines = []
        for i, line in enumerate(lines):
            if (re.search(r'\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}', line) or
                re.search(r'\d{1,2}\s*(?:January|February|March|April|May|June|July|August|September|October|November|December)\s*\d{4}', line, re.IGNORECASE) or
                re.search(r'\d{2}(?:September|December|March|June)\d{4}', line, re.IGNORECASE)):
                date_lines.append((i, line))
        
        # If we found date lines, look for NAV rows after them
        for date_idx, date_line in date_lines:
            # Extract dates from this line
            dates = []
            date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\b|\b\d{2}(?:September|December|March|June)\d{4}\b', date_line, re.IGNORECASE)
            if date_matches:
                dates = date_matches
            else:
                # If no structured dates found, look for month names with years
                month_matches = re.findall(r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s*\d{4}\b|\b\d{2}(?:Sep|Dec|Mar|Jun)[a-z]*\d{4}\b', date_line, re.IGNORECASE)
                if month_matches:
                    dates = month_matches
            
            # If we found dates, look for NAV rows within the next 15 lines
            if dates and len(dates) >= 2:
                print(f"Found date line with dates: {dates}")
                
                # Look for NAV rows after the date line
                for i in range(date_idx + 1, min(date_idx + 15, len(lines))):
                    line = lines[i]
                    
                    # Check if this line might be the NAV row
                    if "net asset value" in line.lower() or "nav" in line.lower().split():
                        print(f"Found potential NAV line: {line}")
                        
                        # Extract all numbers from this line
                        number_matches = re.findall(r"[\d',\.]+", line)
                        values = []
                        
                        for num_str in number_matches:
                            value = clean_number(num_str)
                            if value is not None:
                                values.append(value)
                        
                        # If we found at least 2 values, return them
                        if len(values) >= 2:
                            print(f"Extracted values from text: {values}")
                            return dates[0], values[0], dates[1], values[1]
        
        # If we haven't found anything yet, try another approach
        # Look for "Net asset value" or "NAV" in any line
        for i, line in enumerate(lines):
            if "net asset value" in line.lower() or "nav" in line.lower().split():
                # Found a potential NAV line
                print(f"Found potential NAV line through direct search: {line}")
                
                # Extract numbers from this line
                number_matches = re.findall(r"[\d',\.]+", line)
                values = []
                
                for num_str in number_matches:
                    value = clean_number(num_str)
                    if value is not None:
                        values.append(value)
                
                # If we found at least 2 values
                if len(values) >= 2:
                    # Look for dates in nearby lines
                    date_context = "\n".join(lines[max(0, i-5):min(len(lines), i+5)])
                    date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b|\b\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\b', date_context, re.IGNORECASE)
                    
                    if len(date_matches) >= 2:
                        print(f"Extracted values with dates from context: {values}, {date_matches}")
                        return date_matches[0], values[0], date_matches[1], values[1]
                    else:
                        print(f"Extracted values but no dates found: {values}")
                        return "Period 1", values[0], "Period 2", values[1]
    except Exception as e:
        print(f"Error in direct text scanning: {str(e)}")
    
    return None, None, None, None


def process_pdf_comprehensive(pdf_path):
    """
    Process a PDF using multiple approaches in an intelligent sequence.
    """
    # Step 1: Extract fund name
    fund_name = extract_fund_name(pdf_path)
    print(f"\nProcessing: {os.path.basename(pdf_path)}")
    print(f"Fund name identified: {fund_name}")
    
    # Step 2: Analyze PDF structure to determine best approach
    pdf_info = analyze_pdf_structure(pdf_path)
    
    results = []
    
    # Step 2.5: Special treatment for Partners Group Key Figures format
    if pdf_info["is_partners_group"] and pdf_info["has_key_figures_section"]:
        # Use specialized extraction for Partners Group Key Figures
        period1_label, period1_nav, period2_label, period2_nav = extract_partners_group_key_figures(pdf_path)
        if period1_nav is not None and period2_nav is not None:
            results.append({
                'period1_label': period1_label,
                'period1_nav': period1_nav,
                'period2_label': period2_label,
                'period2_nav': period2_nav,
                'confidence': 0.98,  # Highest confidence for specialized extraction
                'source': 'partners_group_key_figures'
            })
    
    # Step 3: Multi-library text extraction
    text = extract_text_multi_library(pdf_path)
    
    # Step 3.5: Try direct text scanning if we have key figures section
    if pdf_info["has_key_figures_section"] and not results:
        period1_label, period1_nav, period2_label, period2_nav = scan_for_nav_row(text, fund_name)
        if period1_nav is not None and period2_nav is not None:
            results.append({
                'period1_label': period1_label,
                'period1_nav': period1_nav,
                'period2_label': period2_label,
                'period2_nav': period2_nav,
                'confidence': 0.9,  # High confidence for direct text scan
                'source': 'direct_text_scan'
            })
    
    # Step 4: Extract NAV using enhanced patterns from text
    pattern_results = extract_nav_with_enhanced_patterns(text)
    if pattern_results:
        results.extend(pattern_results)
    
    # Step 5: Table-based extraction if the PDF likely has tables
    if pdf_info["has_tables"]:
        tables_data = extract_tables_multi_library(pdf_path)
        table_results = extract_nav_from_tables(tables_data)
        if table_results:
            results.extend(table_results)
    
    # Step 6: If we have Key Figures section, use specialized approach
    if pdf_info["has_key_figures_section"]:
        # Use the original direct_table_extraction function
        period1_label, period1_nav, period2_label, period2_nav = direct_table_extraction(pdf_path)
        if period1_nav is not None and period2_nav is not None:
            results.append({
                'period1_label': period1_label,
                'period1_nav': period1_nav,
                'period2_label': period2_label,
                'period2_nav': period2_nav,
                'confidence': 0.85,  # High confidence for key figures
                'source': 'key_figures'
            })
    
    # Step 7: If all else fails, try the existing fallback methods
    if not results:
        # Use existing methods
        period1_label, period1_nav, period2_label, period2_nav = scan_text_for_nav(pdf_path)
        if period1_nav is not None and period2_nav is not None:
            results.append({
                'period1_label': period1_label,
                'period1_nav': period1_nav,
                'period2_label': period2_label,
                'period2_nav': period2_nav,
                'confidence': 0.6,
                'source': 'scan_text'
            })
        else:
            period1_label, period1_nav, period2_label, period2_nav = last_resort_extraction(pdf_path)
            if period1_nav is not None and period2_nav is not None:
                results.append({
                    'period1_label': period1_label,
                    'period1_nav': period1_nav,
                    'period2_label': period2_label,
                    'period2_nav': period2_nav,
                    'confidence': 0.5,
                    'source': 'last_resort'
                })
    
    # Step 8: Select best result based on confidence
    if results:
        # Sort by confidence
        results.sort(key=lambda x: x['confidence'], reverse=True)
        best_result = results[0]
        
        return {
            'Fund Name': fund_name,
            'Period 1 Label': best_result['period1_label'],
            'Period 1 NAV': best_result['period1_nav'],
            'Period 2 Label': best_result['period2_label'],
            'Period 2 NAV': best_result['period2_nav'],
            'PDF Filename': os.path.basename(pdf_path),
            'Extraction Method': best_result['source'],
            'Confidence': best_result['confidence']
        }
    
    # If all approaches failed
    return {
        'Fund Name': fund_name,
        'Period 1 Label': 'Period 1',
        'Period 1 NAV': None,
        'Period 2 Label': 'Period 2',
        'Period 2 NAV': None,
        'PDF Filename': os.path.basename(pdf_path),
        'Extraction Method': 'failed',
        'Confidence': 0.0
    }


def format_excel(excel_path):
    """
    Apply formatting to the Excel file for better readability.
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Apply header formatting
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format numeric values
        for row in ws.iter_rows(min_row=2):
            for cell in row[2:]:  # Apply only to NAV columns
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
        
        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"Error formatting Excel file: {str(e)}")
        return False


def main():
    """
    Main function to process a folder of PDF files.
    """
    print("=== Comprehensive PDF NAV Extraction Tool ===")
    print("This tool extracts fund names and NAV values from PDF documents using multiple approaches.")
    
    # Folder containing PDF files
    pdf_folder = input("Enter the path to the folder containing PDF files: ")
    
    # Check if the folder exists
    if not os.path.isdir(pdf_folder):
        print(f"Error: Folder '{pdf_folder}' does not exist.")
        return
    
    # Find all PDF files in the folder
    pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
    
    if not pdf_files:
        print(f"No PDF files found in folder '{pdf_folder}'.")
        return
    
    print(f"Found {len(pdf_files)} PDF files. Processing...")
    
    # Process each PDF file using the comprehensive approach
    results = []
    success_count = 0
    failure_count = 0
    
    for i, pdf_path in enumerate(pdf_files):
        print(f"Processing {i+1}/{len(pdf_files)}: {os.path.basename(pdf_path)}...")
        
        # Use the comprehensive approach
        result = process_pdf_comprehensive(pdf_path)
        results.append(result)
        
        if result['Period 1 NAV'] is not None and result['Period 2 NAV'] is not None:
            success_count += 1
            print(f"  ✓ Success! Method: {result['Extraction Method']} (Confidence: {result['Confidence']:.2f})")
            print(f"    {result['Period 1 Label']}: {result['Period 1 NAV']}")
            print(f"    {result['Period 2 Label']}: {result['Period 2 NAV']}")
        else:
            failure_count += 1
            print(f"  ✗ Failed to extract NAV values")
    
    # Create DataFrame from results
    df = pd.DataFrame(results)
    
    # Create output Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(pdf_folder, f"Fund_NAV_Summary_{timestamp}.xlsx")
    
    # Export to Excel
    df.to_excel(output_path, index=False)
    
    # Format the Excel file
    format_excel(output_path)
    
    print(f"\nProcessing complete. Results saved to: {output_path}")
    print(f"Success: {success_count}/{len(pdf_files)} ({success_count/len(pdf_files)*100:.1f}%)")
    print(f"Failed: {failure_count}/{len(pdf_files)} ({failure_count/len(pdf_files)*100:.1f}%)")
    
    # Show extraction method statistics
    method_stats = df['Extraction Method'].value_counts()
    print("\nExtraction Method Statistics:")
    for method, count in method_stats.items():
        print(f"  {method}: {count} ({count/len(pdf_files)*100:.1f}%)")


if __name__ == "__main__":
    main()
