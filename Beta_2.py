import os
import datetime
import re
import pythoncom
import win32com.client
import pandas as pd
import nest_asyncio
import json
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_socketio import SocketIO
from threading import Thread
import openpyxl

# Apply nest_asyncio to allow nested event loops
nest_asyncio.apply()

app = Flask(__name__)
app.secret_key = 'supersecretkey'
socketio = SocketIO(app)

# Load configurations from a JSON file
with open('config.json', 'r') as f:
    config = json.load(f)

DEFAULT_SAVE_PATH = config.get('DEFAULT_SAVE_PATH', 'path_to_default_folder')
LOG_FILE_PATH = config.get('LOG_FILE_PATH', 'logs.txt')
EXCEL_FILE_PATH = config.get('EXCEL_FILE_PATH', 'email_summary.xlsx')

def sanitize_filename(filename):
    allowable_chars = re.compile(r'[^a-zA-Z0-9\s\-\_\.\+\%\(\)]')
    sanitized = allowable_chars.sub('_', filename)
    sanitized = re.sub(r'_+', '_', sanitized)
    sanitized = sanitized.replace(' ', '_')
    return sanitized[:255]

def extract_year_and_month(text, default_year=None):
    date_pattern = re.compile(r"""
    (?i)
    \b(January|February|March|April|May|June|July|August|September|October|November|December)\b|
    \b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b|
    (\d{1,4})\/(\d{1,2})|
    (\d{1,2})[./-](\d{1,2})[./-](\d{2,4})|
    (\d{4})-(\d{1,2})-(\d{1,2})|
    (Q[1-4])['\s]?(\d{2,4})|
    ([1-4]Q)['\s]?(\d{2,4})|
    (\d{2})(\d{4})|
    (\d{4})(\d{2})
    """, re.IGNORECASE | re.VERBOSE)
    match = date_pattern.search(text)

    if match:
        full_month = match.group(1)
        abbr_month = match.group(2)
        numeric_date_1 = match.group(3)
        numeric_date_2 = match.group(4)
        numeric_date_3 = match.group(5)
        numeric_date_4 = match.group(6)
        quarter_1 = match.group(7)
        quarter_year_1 = match.group(8)
        quarter_2 = match.group(9)
        quarter_year_2 = match.group(10)
        month_year_1 = match.group(11)
        year_month_1 = match.group(12)
        year = match.group(13) if match.group(13) else default_year

        # handle full month names
        if full_month:
            month_num = datetime.datetime.strptime(full_month, "%B").strftime("%m")
            month_name = datetime.datetime.strptime(full_month, "%B").strftime("%B")
            return year, f"{month_num}-{month_name}"

        # handle abbr month names
        if abbr_month:
            month_num = datetime.datetime.strptime(abbr_month, "%b").strftime("%m")
            month_name = datetime.datetime.strptime(abbr_month, "%b").strftime("%B")
            return year, f"{month_num}-{month_name}"

        # handle numeric dates
        numeric_dates = [numeric_date_1, numeric_date_2, numeric_date_3, numeric_date_4]
        for numeric_date in numeric_dates:
            if numeric_date:
                date_formats = ["%d.%m.%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%y", "%d/%m/%y", "%d/%m/%Y", 
                                "%m/%d/%y", "%m/%d/%Y", "%m%d%Y", "%d%m%y", "%m%d%y", "%m%d%Y"]
                for date_format in date_formats:
                    try:
                        parsed_date = datetime.datetime.strptime(numeric_date, date_format)
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = year if year else parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        continue

        # handle quarters
        quarter_mappings = {'Q1': '03-March', 'Q2': '06-June', 'Q3': '09-September', 'Q4': '12-December'}
        if quarter_1 and quarter_year_1:
            quarter = quarter_mappings[quarter_1.upper()]
            year = f"20{quarter_year_1}" if len(quarter_year_1) == 2 else quarter_year_1
            return year, quarter

        if quarter_2 and quarter_year_2:
            quarter = quarter_mappings[quarter_2.upper()]
            year = f"20{quarter_year_2}" if len(quarter_year_2) == 2 else quarter_year_2
            return year, quarter

        # handle month-year or year-month formats like 052024 or 202405
        if month_year_1 and len(month_year_1) == 2:
            month_num = month_year_1
            month_name = datetime.datetime.strptime(month_num, "%m").strftime("%B")
            year = year_month_1 if year_month_1 else default_year
            return year, f"{month_num}-{month_name}"

        if year_month_1 and len(year_month_1) == 4:
            year = year_month_1
            month_num = match.group(12)
            month_name = datetime.datetime.strptime(month_num, "%m").strftime("%B")
            return year, f"{month_num}-{month_name}"

    return default_year, None

def extract_year_for_keywords(text):
    year_pattern = re.compile(r"\b(20\d{2})\b")
    match = year_pattern.search(text)
    if match:
        return match.group(1)
    return None

def find_path_for_sender(sender, subject, sender_path_table):
    rows = sender_path_table[sender_path_table['sender'].str.lower() == sender.lower()]

    for _, row in rows.iterrows():
        keywords = str(row.get('keywords', '')).split(';')
        for keyword in keywords:
            if keyword.lower() in subject.lower():
                return row['keyword_path'], row['special_case'], True

    for _, row in rows.iterrows():
        if pd.notna(row['coper_name']) and row['coper_name'].lower() in subject.lower():
            return row['save_path'], row['special_case'], False

    if not rows.empty():
        return rows.iloc[0]['save_path'], rows.iloc[0]['special_case'], False

    return None, None, False

def update_excel_summary(date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails):
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Summary'
        sheet.append(['Date', 'Total Emails', 'Saved in Default', 'Saved in Actual Paths', 'Not Saved'])

    sheet = workbook.active
    sheet.append([date_str, total_emails, saved_default, saved_actual, not_saved])

    if 'Failed Emails' not in workbook.sheetnames:
        failed_sheet = workbook.create_sheet('Failed Emails')
        failed_sheet.append(['Date', 'Email Address', 'Subject'])
    else:
        failed_sheet = workbook['Failed Emails']

    for email in failed_emails:
        failed_sheet.append([date_str, email['email_address'], email['subject']])

    workbook.save(EXCEL_FILE_PATH)

def save_email(item, save_path, special_case):
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    
    valid_extensions = ('.xlsx', '.xls', '.csv', '.pdf', '.doc', '.docx')
    if special_case and special_case.lower() == 'yes' and item.Attachments.Count > 0:
        for attachment in item.Attachments:
            # Only consider attachments with specific file types
            if attachment.FileName.lower().endswith(valid_extensions):
                filename = f"{sanitize_filename(attachment.Filename)}.msg"
                break
        else:
            # If no valid attachment is found, fallback to using the subject
            filename = f"{sanitize_filename(item.Subject)}.msg"
    else:
        filename = f"{sanitize_filename(item.Subject)}.msg"

    item.SaveAs(os.path.join(save_path, filename), 3)
    return filename

def process_email(item, sender_path_table, default_year, specific_date_str):
    logs = []
    failed_emails = []
    retries = 3
    processed = False
    while retries > 0 and not processed:
        try:
            sender_email = item.SenderEmailAddress.lower() if hasattr(item, 'SenderEmailAddress') else item.Sender.Address.lower()
            year, month = extract_year_and_month(item.Subject, default_year)
            if not year or not month:
                for attachment in item.Attachments:
                    year, month = extract_year_and_month(attachment.Filename, default_year)
                    if year and month:
                        break
            if not year:
                year = default_year
            base_path, special_case, is_keyword_path = find_path_for_sender(sender_email, item.Subject, sender_path_table)
            if base_path:
                save_path = os.path.join(base_path, str(year), month if month else '')
            else:
                save_path = os.path.join(DEFAULT_SAVE_PATH, specific_date_str)

            filename = save_email(item, save_path, special_case)
            logs.append(f"Saved: {filename} to {save_path}")
            processed = True
        except pythoncom.com_error as com_err:
            retries -= 1
            logs.append(f"COM Error handling email with subject '{item.Subject}' (Code: {com_err.args})")
            if retries == 0:
                logs.append(f"Failed to save the email '{item.Subject}' after 3 retries")
                failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
        except Exception as e:
            retries = 0
            logs.append(f"Error handling email with subject '{item.Subject}': {str(e)}")
            failed_emails.append({'email_address': sender_email, 'subject': item.Subject})

    return logs, failed_emails

def save_emails_from_senders_on_date(email_address, specific_date_str, sender_path_table, default_year):
    logs = []
    pythoncom.CoInitialize()
    specific_date = datetime.datetime.strptime(specific_date_str, '%Y-%m-%d').date()
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = None

    for store in outlook.Stores:
        if store.DisplayName.lower() == email_address.lower() or store.ExchangeStoreType == 3:
            try:
                root_folder = store.GetRootFolder()
                for folder in root_folder.Folders:
                    if folder.Name.lower() == "inbox":
                        inbox = folder
                        break
                if inbox is not None:
                    break
            except AttributeError as e:
                logs.append(f"Error accessing inbox: {str(e)}")
                continue

    if inbox is None:
        logs.append(f"No Inbox found for the account with email address: {email_address}")
        pythoncom.CoUninitialize()
        with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
            f.writelines("\n".join(logs))
        return

    items = inbox.Items
    items.Sort("[ReceivedTime]", True)
    items = items.Restrict(f"[ReceivedTime] >= '{specific_date.strftime('%m/%d/%Y')} 00:00 AM' AND [ReceivedTime] <= '{specific_date.strftime('%m/%d/%Y')} 11:59 PM'")

    total_emails = 0
    saved_default = 0
    saved_actual = 0
    not_saved = 0
    failed_emails = []

    for item in items:
        total_emails += 1
        email_logs, email_failed_emails = process_email(item, sender_path_table, default_year, specific_date_str)
        logs.extend(email_logs)
        failed_emails.extend(email_failed_emails)
        if any(DEFAULT_SAVE_PATH in log for log in email_logs):
            saved_default += 1
        else:
            saved_actual += 1

    pythoncom.CoUninitialize()
    with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
        f.writelines("\n".join(logs))

    update_excel_summary(specific_date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        date_str = request.form['date']
        default_year = request.form['default_year']
        file = request.files['file']
        if file and date_str and default_year:
            try:
                datetime.datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                flash("Invalid date format. Please enter the date in YYYY-MM-DD format.", 'error')
                return redirect(url_for('index'))

            if not (default_year.isdigit() and len(default_year) == 4):
                flash("Invalid year format. Please enter the year in YYYY format.", 'error')
                return redirect(url_for('index'))

            filename = file.filename
            filepath = os.path.join('uploads', filename)
            file.save(filepath)
            sender_path_table = pd.read_csv(filepath)

            account_email_address = "hf_data@bofa.com"
            socketio.start_background_task(save_emails_from_senders_on_date, account_email_address, date_str, sender_path_table, default_year)
            return redirect(url_for('results'))

    return render_template('index.html')

@app.route('/results')
def results():
    logs = []
    if os.path.exists(LOG_FILE_PATH):
        with open(LOG_FILE_PATH, 'r') as f:
            logs = f.readlines()

    return render_template('results.html', logs=logs)

def run_app():
    socketio.run(app, debug=True, use_reloader=False, allow_unsafe_werkzeug=True)

if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    if not os.path.exists(DEFAULT_SAVE_PATH):
        os.makedirs(DEFAULT_SAVE_PATH)

    thread = Thread(target=run_app)
    thread.start()
