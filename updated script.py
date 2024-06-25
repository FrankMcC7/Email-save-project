import os
import datetime
import re
import pythoncom
import win32com.client
import pandas as pd
import nest_asyncio
import json
from flask import Flask, render_template, request, redirect, url_for
from flask_socketio import SocketIO
from threading import Thread
import openpyxl

# Apply nest_asyncio to allow nested event loops
nest_asyncio.apply()

app = Flask(__name__)
app.secret_key = 'supersecretkey'
socketio = SocketIO(app)

# Load configurations from a JSON file
with open('config.json', 'r') as f:
    config = json.load(f)

DEFAULT_SAVE_PATH = config.get('DEFAULT_SAVE_PATH', 'path_to_default_folder')
LOG_FILE_PATH = config.get('LOG_FILE_PATH', 'logs.txt')
EXCEL_FILE_PATH = config.get('EXCEL_FILE_PATH', 'email_summary.xlsx')

def sanitize_filename(filename):
    allowable_chars = re.compile(r'[^a-zA-Z0-9\s\-\_\.\+\%\(\)]')
    sanitized = allowable_chars.sub('_', filename)
    sanitized = re.sub(r'_+', '_', sanitized)
    sanitized = sanitized.replace(' ', '_')
    return sanitized[:255]

def extract_year_and_month(text, default_year=None):
    date_pattern = re.compile(r"""
        (?:
            \b(January|February|March|April|May|June|July|August|September|October|November|December)\b|
            \b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b|
            \b(Q[1-4])\b|
            (\d{1,2}[./-]\d{1,2}[./-]\d{2,4})|
            (\d{1,2}[./-]\d{2,4})|
            (\d{1,2}\s\d{1,2}\s\d{2,4})|
            (\d{1,2}\s\d{4})
        )
        (?:\s*(\d{4}))?
    """, re.IGNORECASE | re.VERBOSE)
    
    match = date_pattern.search(text)
    if match:
        full_month = match.group(1)
        abbr_month = match.group(2)
        quarter = match.group(3)
        numeric_date = match.group(4) or match.group(5)
        space_separated_date = match.group(6)
        space_separated_date_without_day = match.group(7)
        year = match.group(8) if match.group(8) else default_year

        if full_month:
            month_num = datetime.datetime.strptime(full_month, '%B').strftime('%m')
            month_name = datetime.datetime.strptime(full_month, '%B').strftime('%B')
            return year, f"{month_num}-{month_name}"

        if abbr_month:
            month_num = datetime.datetime.strptime(abbr_month, '%b').strftime('%m')
            month_name = datetime.datetime.strptime(abbr_month, '%b').strftime('%B')
            return year, f"{month_num}-{month_name}"

        if quarter:
            month_map = {'Q1': '03-March', 'Q2': '06-June', 'Q3': '09-September', 'Q4': '12-December'}
            month = month_map[quarter]
            return year, month

        if numeric_date:
            date_formats = ["%d.%m.%y", "%m.%d.%y", "%d.%m.%Y", "%m.%d.%Y", "%d/%m/%y", "%m/%d/%y", "%d/%m/%Y", "%m/%d/%Y", "%d%m%y", "%m%d%y", "%d%m%Y", "%m%d%Y"]
            for date_format in date_formats:
                try:
                    parsed_date = datetime.datetime.strptime(numeric_date, date_format)
                    month_num = parsed_date.strftime('%m')
                    month_name = parsed_date.strftime('%B')
                    year = year if year else parsed_date.strftime('%Y')
                    return year, f"{month_num}-{month_name}"
                except ValueError:
                    continue

        if space_separated_date:
            try:
                parsed_date = datetime.datetime.strptime(space_separated_date, "%m %d %Y")
                month_num = parsed_date.strftime('%m')
                month_name = parsed_date.strftime('%B')
                year = year if year else parsed_date.strftime('%Y')
                return year, f"{month_num}-{month_name}"
            except ValueError:
                pass

        if space_separated_date_without_day:
            try:
                parsed_date = datetime.datetime.strptime(space_separated_date_without_day, "%m %Y")
                month_num = parsed_date.strftime('%m')
                month_name = parsed_date.strftime('%B')
                year = year if year else parsed_date.strftime('%Y')
                return year, f"{month_num}-{month_name}"
            except ValueError:
                pass

    return default_year, None

def update_excel_summary(date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails):
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    else:
        workbook = openpyxl.Workbook()
        workbook.active.append(['Date', 'Total Emails', 'Saved in Default', 'Saved in Actual Paths', 'Not Saved'])

    sheet = workbook.active
    sheet.append([date_str, total_emails, saved_default, saved_actual, not_saved])

    if 'Failed Emails' not in workbook.sheetnames:
        workbook.create_sheet('Failed Emails')

    failed_sheet = workbook['Failed Emails']
    if failed_sheet.max_row == 1:
        failed_sheet.append(['Date', 'Email Address', 'Subject'])

    for email in failed_emails:
        failed_sheet.append([date_str, email['email_address'], email['subject']])

    workbook.save(EXCEL_FILE_PATH)

def find_path_for_sender(sender, subject, sender_path_table):
    rows = sender_path_table[sender_path_table['sender'].str.lower() == sender.lower()]
    if len(rows) > 1:
        for _, row in rows.iterrows():
            if pd.notna(row['coper_name']) and row['coper_name'].lower() in subject.lower():
                return row['save_path'], row['keyword_path'], row['keywords']
    if not rows.empty:
        row = rows.iloc[0]
        return row['save_path'], row['keyword_path'], row['keywords']
    return None, None, None

def save_emails_from_senders_on_date(email_address, specific_date_str, sender_path_table, default_year):
    logs = []
    pythoncom.CoInitialize()
    specific_date = datetime.datetime.strptime(specific_date_str, '%Y-%m-%d').date()
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = None

    for store in outlook.Stores:
        if store.DisplayName.lower() == email_address.lower() or store.ExchangeStoreType == 3:
            try:
                root_folder = store.GetRootFolder()
                for folder in root_folder.Folders:
                    if folder.Name.lower() == "inbox":
                        inbox = folder
                        break
                if inbox is not None:
                    break
            except AttributeError as e:
                logs.append(f"Error accessing inbox: {str(e)}")
                continue

    if inbox is None:
        logs.append(f"No Inbox found for the account with the email address: {email_address}")
        pythoncom.CoUninitialize()
        with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
            for log in logs:
                f.write(f"{log}\n")
        return

    items = inbox.Items
    items.Sort("[ReceivedTime]", True)
    items = items.Restrict(f"[ReceivedTime] >= '{specific_date.strftime('%m/%d/%Y')} 00:00 AM' AND [ReceivedTime] <= '{specific_date.strftime('%m/%d/%Y')} 11:59 PM'")

    total_emails = 0
    saved_default = 0
    saved_actual = 0
    not_saved = 0
    failed_emails = []

    max_retries = 3

    for item in items:
        total_emails += 1
        retries = 0
        processed = False

        if hasattr(item, 'SenderEmailAddress') or hasattr(item, 'Sender'):
            sender_email = item.SenderEmailAddress.lower() if hasattr(item, 'SenderEmailAddress') else item.Sender.Address.lower()
            save_path, keyword_path, keywords = find_path_for_sender(sender_email, item.Subject, sender_path_table)

        while retries < max_retries and not processed:
            try:
                if not sender_email:
                    logs.append(f"Error: Email item has no sender address.")
                    failed_emails.append({'email_address': 'Unknown', 'subject': item.Subject})
                    not_saved += 1
                    break

                if not save_path:
                    year, month = extract_year_and_month(item.Subject, default_year)
                    year_month_path = os.path.join(DEFAULT_SAVE_PATH, sender_email, year, month if month else "")
                    if not os.path.exists(year_month_path):
                        os.makedirs(year_month_path)
                    subject = sanitize_filename(item.Subject)
                    filename = f"{subject}.msg"
                    try:
                        item.SaveAs(os.path.join(year_month_path, filename), 3)
                        logs.append(f"Saved: {filename} to {year_month_path}")
                        saved_default += 1
                        processed = True
                    except Exception as save_err:
                        logs.append(f"Failed to save email to default path: {str(save_err)}")
                        failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
                        not_saved += 1
                    continue

                keywords_list = [kw.strip().lower() for kw in keywords.split(';')] if keywords else []
                keyword_matched = any(keyword in item.Subject.lower() for keyword in keywords_list)

                if not keyword_matched and item.Attachments.Count > 0:
                    for attachment in item.Attachments:
                        if any(keyword in attachment.FileName.lower() for keyword in keywords_list):
                            keyword_matched = True
                            break

                if keyword_matched:
                    subject = sanitize_filename(item.Subject)
                    filename = f"{subject}_{specific_date_str}.msg"
                    try:
                        item.SaveAs(os.path.join(keyword_path, filename), 3)
                        logs.append(f"Saved keyword case: {filename} to {keyword_path}")
                        saved_actual += 1
                        processed = True
                    except Exception as save_err:
                        logs.append(f"Failed to save keyword case email: {str(save_err)}")
                        failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
                        not_saved += 1
                else:
                    year, month = extract_year_and_month(item.Subject, default_year)
                    year_month_path = os.path.join(save_path, year, month if month else "")
                    if not os.path.exists(year_month_path):
                        os.makedirs(year_month_path)
                    subject = sanitize_filename(item.Subject)
                    filename = f"{subject}.msg"
                    try:
                        item.SaveAs(os.path.join(year_month_path, filename), 3)
                        logs.append(f"Saved: {filename} to {year_month_path}")
                        saved_actual += 1
                        processed = True
                    except Exception as save_err:
                        logs.append(f"Failed to save email: {str(save_err)}")
                        failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
                        not_saved += 1

            except pythoncom.com_error as com_err:
                error_code, _, error_message, _ = com_err.args
                logs.append(f"COM Error handling email with subject '{item.Subject}': {error_message} (Code: {error_code})")
                retries += 1
                if retries >= max_retries:
                    logs.append(f"Failed to save email '{item.Subject}' after {max_retries} retries.")
                    failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
                    not_saved += 1
            except Exception as e:
                logs.append(f"Error handling email with subject '{item.Subject}': {str(e)}")
                failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
                not_saved += 1
                retries = max_retries

        if not processed:
            year, month = extract_year_and_month(item.Subject, default_year)
            year_month_path = os.path.join(DEFAULT_SAVE_PATH, sender_email, year, month if month else "")
            if not os.path.exists(year_month_path):
                os.makedirs(year_month_path)
            subject = sanitize_filename(item.Subject)
            filename = f"{subject}.msg"
            try:
                item.SaveAs(os.path.join(year_month_path, filename), 3)
                logs.append(f"Saved: {filename} to {year_month_path}")
                saved_default += 1
            except Exception as save_err:
                logs.append(f"Failed to save email to default path: {str(save_err)}")
                failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
                not_saved += 1

    pythoncom.CoUninitialize()
    with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
        for log in logs:
            f.write(f"{log}\n")

    update_excel_summary(specific_date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        date_str = request.form['date']
        default_year = request.form['default_year']
        file = request.files['file']
        if file and date_str and default_year:
            try:
                datetime.datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                return "Invalid date format. Please enter the date in YYYY-MM-DD format."
            
            if not (default_year.isdigit() and len(default_year) == 4):
                return "Invalid year format. Please enter the year in YYYY format."

            filename = file.filename
            filepath = os.path.join('uploads', filename)
            file.save(filepath)
            sender_path_table = pd.read_csv(filepath)

            account_email_address = 'your_account_email@example.com'
            socketio.start_background_task(save_emails_from_senders_on_date, account_email_address, date_str, sender_path_table, default_year)
            return redirect(url_for('results'))

    return render_template('index.html')

@app.route('/results')
def results():
    logs = []
    if os.path.exists(LOG_FILE_PATH):
        with open(LOG_FILE_PATH, 'r', encoding='utf-8') as f:
            logs = f.readlines()
    return render_template('results.html', logs=logs)

def run_app():
    socketio.run(app, debug=True, use_reloader=False, allow_unsafe_werkzeug=True)

if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    if not os.path.exists(DEFAULT_SAVE_PATH):
        os.makedirs(DEFAULT_SAVE_PATH)
    
    thread = Thread(target=run_app)
    thread.start()
