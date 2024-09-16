import os
import datetime
import re
import pythoncom
import win32com.client
import pandas as pd
import json
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_socketio import SocketIO
import openpyxl
import unicodedata

app = Flask(__name__)
app.secret_key = 'supersecretkey'
socketio = SocketIO(app)

# Load configurations from a JSON file
try:
    with open('config.json', 'r') as f:
        config = json.load(f)
except FileNotFoundError:
    config = {}
    print("Configuration file 'config.json' not found. Using default settings.")

DEFAULT_SAVE_PATH = config.get('DEFAULT_SAVE_PATH', 'path_to_default_folder')
LOG_FILE_PATH = config.get('LOG_FILE_PATH', 'logs.txt')
EXCEL_FILE_PATH = config.get('EXCEL_FILE_PATH', 'email_summary.xlsx')

def sanitize_filename(filename):
    # Normalize unicode characters to their closest ASCII equivalent (e.g., é -> e)
    normalized_filename = unicodedata.normalize('NFKD', filename).encode('ASCII', 'ignore').decode('ASCII')

    # Replace common problematic characters with underscores or remove them
    sanitized = re.sub(r'[<>:"/\\|?*\[\]\'`~!@#$%^&*()+={};,]', '_', normalized_filename)

    # Replace dots (.) followed by a space or end of the string with an underscore, except for file extensions
    sanitized = re.sub(r'\.(?=\s|$)', '_', sanitized)

    # Replace any sequence of multiple special characters (like "--") with a single underscore
    sanitized = re.sub(r'_+', '_', sanitized)

    # Trim leading and trailing underscores or spaces
    sanitized = sanitized.strip(' _')

    # Limit filename length to avoid filesystem issues (255 characters max, considering extension)
    sanitized = sanitized[:255]

    return sanitized

def extract_date_from_text(text, default_year=None):
    import re
    import datetime

    # Map quarters to months
    quarter_mappings = {
        '1': '03-March', '2': '06-June', '3': '09-September', '4': '12-December',
        'Q1': '03-March', 'Q2': '06-June', 'Q3': '09-September', 'Q4': '12-December',
    }

    # Prepare text by replacing various separators with spaces
    text = text.replace("'", "").replace(",", " ").replace("-", " ").replace("/", " ").replace(".", " ")
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()  # Remove leading/trailing whitespace

    # Define regex patterns to extract date components
    patterns = [
        # Match full or abbreviated month names with optional day and year
        r'(?i)\s*(?:\b(?:on|as of|for)\b\s*)?'        # Optional prefixes with word boundaries
        r'(?:\b(\d{1,2})\b\s*)?'                      # Optional day with word boundary
        r'\b('                                        # Word boundary before month name
        r'January|February|March|April|May|June|'
        r'July|August|September|October|November|December|'
        r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec'
        r')\b'                                        # Word boundary after month name
        r'[\s,]*'                                     # Optional separator
        r'(\d{4}|\d{2})?\b',                          # Optional year with word boundary

        # Pattern for compact dates with 6 or 8 digits
        r'\b(\d{6,8})\b',

        # Pattern for dates with separators: DD/MM/YY, DD-MM-YYYY, etc.
        r'\b(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\b',

        # Match year and month: YYYY MM or MM YYYY
        r'\b(\d{4})\s+(\d{1,2})\b',
        r'\b(\d{1,2})\s+(\d{4})\b',

        # Match month and year without separator: MMYYYY or YYYYMM
        r'\b(\d{2})(\d{4})\b',
        r'\b(\d{4})(\d{2})\b',

        # Match quarters: Q1 2024 or 1Q 2024
        r'\b(Q[1-4]|[1-4]Q)[\s]*(\d{2,4})\b',
    ]

    for pattern in patterns:
        matches = re.findall(pattern, text)
        for match in matches:
            # Handle month name patterns
            if len(match) == 3 and re.match(
                r'(?i)^(January|February|March|April|May|June|July|August|September|October|November|December|'
                r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)$',
                match[1]
            ):
                day, month_str, year = match
                day = day.strip() if day else '1'  # Default to first day if not provided
                year = year.strip() if year else default_year
                if not year:
                    continue
                # Handle two-digit years
                if len(year) == 2:
                    year = '20' + year
                date_str = f"{day} {month_str} {year}"
                date_formats = ['%d %B %Y', '%d %b %Y']
                # Try parsing the date string with the date formats
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.datetime.strptime(date_str, fmt)
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        continue

            # Handle compact date patterns with 6 or 8 digits
            elif len(match) == 1 and match[0].isdigit():
                date_str = match[0]
                if len(date_str) == 6:
                    date_formats = ['%d%m%y', '%m%d%y', '%y%m%d', '%y%d%m']
                elif len(date_str) == 8:
                    date_formats = ['%d%m%Y', '%m%d%Y', '%Y%m%d', '%Y%d%m']
                else:
                    continue
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.datetime.strptime(date_str, fmt)
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        continue

            # Handle dates with separators: DD/MM/YY, DD-MM-YYYY, etc.
            elif len(match) == 3 and all(part.isdigit() for part in match):
                part1, part2, part3 = match
                combinations = [
                    (f"{part1} {part2} {part3}", ['%d %m %Y', '%d %m %y']),
                    (f"{part2} {part1} {part3}", ['%m %d %Y', '%m %d %y']),
                    (f"{part3} {part2} {part1}", ['%Y %m %d', '%y %m %d']),
                ]
                parsed = False
                for date_str, date_formats in combinations:
                    for fmt in date_formats:
                        try:
                            parsed_date = datetime.datetime.strptime(date_str, fmt)
                            month_num = parsed_date.strftime('%m')
                            month_name = parsed_date.strftime('%B')
                            year = parsed_date.strftime('%Y')
                            return year, f"{month_num}-{month_name}"
                        except ValueError:
                            continue
                    if parsed:
                        break
                if not parsed:
                    continue

            # Handle other patterns as before...
            elif len(match) == 2 and match[0].isdigit() and match[1].isdigit():
                num1, num2 = match
                if len(num1) == 4 and len(num2) <= 2:
                    date_str = f"{num1} {num2}"
                    date_formats = ['%Y %m']
                elif len(num2) == 4 and len(num1) <= 2:
                    date_str = f"{num1} {num2}"
                    date_formats = ['%m %Y']
                else:
                    continue
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.datetime.strptime(date_str, fmt)
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        continue

            # Handle month and year without separator: MMYYYY or YYYYMM
            elif len(match) == 2 and match[0].isdigit() and match[1].isdigit():
                part1, part2 = match
                if len(part1) == 2 and len(part2) == 4:
                    date_str = f"{part1} {part2}"
                    date_formats = ['%m %Y']
                elif len(part1) == 4 and len(part2) == 2:
                    date_str = f"{part1} {part2}"
                    date_formats = ['%Y %m']
                else:
                    continue
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.datetime.strptime(date_str, fmt)
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        continue

            # Handle quarter patterns
            elif len(match) == 2:
                quarter_str, year = match
                quarter = re.sub(r'[^1-4]', '', quarter_str)
                if not quarter:
                    continue
                year = year.strip()
                if len(year) == 2:
                    year = '20' + year
                if year and quarter in quarter_mappings:
                    return year, quarter_mappings[quarter]
                else:
                    continue
            else:
                continue

    # If no date is found, return default_year and None
    return default_year, None

def find_save_path(sender, subject, sender_path_table):
    # Check if the sender exists in the CSV file
    rows = sender_path_table[sender_path_table['sender'].str.lower() == sender.lower()]

    # If the sender is not found in the CSV, treat it as a default path email
    if rows.empty:
        return None, None, False  # Indicate that this is a default path email

    # If the sender email has multiple entries, apply coper_name matching logic
    if len(rows) > 1:
        for _, row in rows.iterrows():
            coper_name = str(row.get('coper_name', '')).strip().lower()
            if coper_name and coper_name in subject.lower():  # Match the coper_name in subject
                # Check if the subject also contains any keywords associated with this coper_name
                keywords = str(row.get('keywords', '')).split(';')
                for keyword in keywords:
                    if keyword.lower() in subject.lower():
                        keyword_path = row.get('keyword_path', '')
                        return keyword_path, False, True  # Save in keyword path
                save_path = row.get('save_path', '')
                return save_path, False, True  # Save in save path if no keywords match
        # If no coper_name matches, return None to indicate default save path
        return None, None, False  # Indicate that this is a default path email

    # If the sender email has a unique entry, apply normal keyword and special case logic
    else:
        row = rows.iloc[0]  # Since it's a unique entry, take the first (and only) row
        keywords = str(row.get('keywords', '')).split(';')
        for keyword in keywords:
            if keyword.lower() in subject.lower():
                keyword_path = row.get('keyword_path', '')
                return keyword_path, False, True  # Save in keyword path
        # Check if it's a special case
        special_case_value = str(row.get('special_case', '')).strip().lower() == 'yes'
        save_path = row.get('save_path', '')
        return save_path, special_case_value, True  # Save in save path or special case path

def update_excel_summary(date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails):
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Summary'
        sheet.append(['Date', 'Total Emails', 'Saved in Default', 'Saved in Actual Paths', 'Not Saved'])

    sheet = workbook.active
    sheet.append([date_str, total_emails, saved_default, saved_actual, not_saved])

    if 'Failed Emails' not in workbook.sheetnames:
        failed_sheet = workbook.create_sheet('Failed Emails')
        failed_sheet.append(['Date', 'Email Address', 'Subject'])
    else:
        failed_sheet = workbook['Failed Emails']

    for email in failed_emails:
        failed_sheet.append([date_str, email['email_address'], email['subject']])

    workbook.save(EXCEL_FILE_PATH)

def save_email(item, save_path, special_case):
    try:
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        
        valid_extensions = ('.xlsx', '.xls', '.csv', '.pdf', '.doc', '.docx')
        if special_case and item.Attachments.Count > 0:
            for attachment in item.Attachments:
                if attachment.FileName.lower().endswith(valid_extensions):
                    filename_base = sanitize_filename(os.path.splitext(attachment.FileName)[0])
                    break
            else:
                filename_base = sanitize_filename(item.Subject)
        else:
            filename_base = sanitize_filename(item.Subject)
        
        # Ensure the filename does not exceed the maximum length
        extension = ".msg"
        max_filename_length = 255 - len(save_path) - len(extension) - 1
        if len(filename_base) > max_filename_length:
            filename_base = filename_base[:max_filename_length]
        
        filename = f"{filename_base}{extension}"
        full_path = os.path.join(save_path, filename)
        
        # Check if a file with the same name already exists
        counter = 1
        while os.path.exists(full_path):
            # If it exists, add a suffix to the filename
            filename = f"{filename_base}_{counter}{extension}"
            full_path = os.path.join(save_path, filename)
            counter += 1
        
        # Attempt to save the email
        item.SaveAs(full_path, 3)
        return filename
    except pythoncom.com_error as com_err:
        # Log more details in case of an error
        error_message = f"COM Error saving email '{item.Subject}' to '{save_path}': {str(com_err)}"
        print(error_message)
        raise  # Re-raise the error after logging
    except Exception as e:
        # Log general exceptions
        error_message = f"General Error saving email '{item.Subject}' to '{save_path}': {str(e)}"
        print(error_message)
        raise  # Re-raise the error after logging

def process_email(item, sender_path_table, default_year, specific_date_str):
    logs = []
    failed_emails = []
    retries = 3
    processed = False

    while retries > 0 and not processed:
        try:
            sender_email = item.SenderEmailAddress.lower() if hasattr(item, 'SenderEmailAddress') else item.Sender.Address.lower()
            year, month = extract_date_from_text(item.Subject, default_year)
            if not year or not month:
                for attachment in item.Attachments:
                    year, month = extract_date_from_text(attachment.FileName, default_year)
                    if year and month:
                        break
            year = year or default_year

            # Determine the save path based on the sender and subject
            base_path, special_case, is_csv_path = find_save_path(sender_email, item.Subject, sender_path_table)

            # Ensure base_path is not None
            if base_path is None:
                base_path = DEFAULT_SAVE_PATH

            # If it's a default path email (sender not in CSV)
            if not is_csv_path:
                save_path = os.path.join(base_path, specific_date_str)
            else:
                # Sender is in CSV
                if special_case:
                    # Special case emails are saved in base_path/year
                    save_path = os.path.join(base_path, str(year))
                else:
                    # Normal case: Save in year/month folder if both are identified
                    if month:
                        save_path = os.path.join(base_path, str(year), month)
                    else:
                        save_path = os.path.join(base_path, str(year))

            # Debugging statements to verify the save path
            print(f"Email from: {sender_email}")
            print(f"Subject: {item.Subject}")
            print(f"Special Case: {special_case}")
            print(f"Save Path: {save_path}")

            filename = save_email(item, save_path, special_case)
            logs.append(f"Saved: {filename} to {save_path}")
            processed = True
        except pythoncom.com_error as com_err:
            retries -= 1
            logs.append(f"COM Error handling email '{item.Subject}' from '{sender_email}' (Code: {com_err.args})")
            if retries == 0:
                logs.append(f"Failed to save the email '{item.Subject}' from '{sender_email}' after 3 retries")
                failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
        except Exception as e:
            retries = 0
            logs.append(f"Error handling email '{item.Subject}': {str(e)}")
            failed_emails.append({'email_address': sender_email, 'subject': item.Subject})

    return logs, failed_emails

def save_emails_from_senders_on_date(email_address, specific_date_str, sender_path_table, default_year):
    logs = []
    pythoncom.CoInitialize()
    specific_date = datetime.datetime.strptime(specific_date_str, '%Y-%m-%d').date()
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

    inbox = None
    for store in outlook.Stores:
        if store.DisplayName.lower() == email_address.lower() or store.ExchangeStoreType == 3:
            try:
                root_folder = store.GetRootFolder()
                inbox = next((folder for folder in root_folder.Folders if folder.Name.lower() == "inbox"), None)
                if inbox:
                    break
            except AttributeError as e:
                logs.append(f"Error accessing inbox: {str(e)}")
                continue

    if not inbox:
        logs.append(f"No Inbox found for the account with email address: {email_address}")
        pythoncom.CoUninitialize()
        with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
            f.writelines("\n".join(logs))
        return

    items = inbox.Items
    items.Sort("[ReceivedTime]", True)
    items = items.Restrict(f"[ReceivedTime] >= '{specific_date.strftime('%m/%d/%Y')} 00:00 AM' AND [ReceivedTime] <= '{specific_date.strftime('%m/%d/%Y')} 11:59 PM'")

    total_emails, saved_default, saved_actual, not_saved = 0, 0, 0, 0
    failed_emails = []

    for item in items:
        total_emails += 1
        email_logs, email_failed_emails = process_email(item, sender_path_table, default_year, specific_date_str)
        logs.extend(email_logs)
        failed_emails.extend(email_failed_emails)
        if any(DEFAULT_SAVE_PATH in log for log in email_logs):
            saved_default += 1
        else:
            saved_actual += 1

    pythoncom.CoUninitialize()
    with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
        f.writelines("\n".join(logs))

    update_excel_summary(specific_date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        date_str = request.form['date']
        default_year = request.form['default_year']
        file = request.files['file']
        if file and date_str and default_year:
            try:
                datetime.datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                flash("Invalid date format. Please enter the date in YYYY-MM-DD format.", 'error')
                return redirect(url_for('index'))

            if not (default_year.isdigit() and len(default_year) == 4):
                flash("Invalid year format. Please enter the year in YYYY format.", 'error')
                return redirect(url_for('index'))

            filepath = os.path.join('uploads', sanitize_filename(file.filename))
            file.save(filepath)

            try:
                sender_path_table = pd.read_csv(filepath, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    sender_path_table = pd.read_csv(filepath, encoding='latin1')
                except Exception as e:
                    flash("Error reading the CSV file. Please ensure it's properly formatted.", 'error')
                    return redirect(url_for('index'))

            # Standardize column names to lowercase
            sender_path_table.columns = sender_path_table.columns.str.lower()

            account_email_address = "hf_data@bofa.com"
            socketio.start_background_task(save_emails_from_senders_on_date, account_email_address, date_str, sender_path_table, default_year)
            return redirect(url_for('results'))

    return render_template('index.html')

@app.route('/results')
def results():
    logs = []
    if os.path.exists(LOG_FILE_PATH):
        with open(LOG_FILE_PATH, 'r', encoding='utf-8') as f:
            logs = f.readlines()

    return render_template('results.html', logs=logs)

if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    os.makedirs(DEFAULT_SAVE_PATH, exist_ok=True)
    socketio.run(app, debug=True, use_reloader=False)
