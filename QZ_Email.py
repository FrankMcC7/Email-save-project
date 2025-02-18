import os
import datetime
import re
import pythoncom
import win32com.client
import pandas as pd
import openpyxl
import unicodedata

# Hardcoded paths and settings
DEFAULT_SAVE_PATH = 'path_to_default_folder'
LOG_FILE_PATH = 'logs.txt'
EXCEL_FILE_PATH = 'email_summary.xlsx'
SENDER_PATH_TABLE_PATH = r'path_to_sender_path_table.csv'  # Update this to your CSV file path

def sanitize_filename(filename):
    # Normalize unicode characters to their closest ASCII equivalent (e.g., é -> e)
    normalized_filename = unicodedata.normalize('NFKD', filename).encode('ASCII', 'ignore').decode('ASCII')
    # Replace common problematic characters with underscores or remove them
    sanitized = re.sub(r'[<>:"/\\|?*\[\]\'`~!@#$%^&*()+={};,]', '_', normalized_filename)
    # Replace dots (.) followed by a space or end of the string with an underscore
    sanitized = re.sub(r'\.(?=\s|$)', '_', sanitized)
    # Replace multiple underscores with a single underscore
    sanitized = re.sub(r'_+', '_', sanitized)
    # Trim leading and trailing underscores or spaces
    sanitized = sanitized.strip(' _')
    # Limit filename length
    sanitized = sanitized[:255]
    return sanitized

def extract_date_from_text(text, default_year=None):
    # Map quarters to months
    quarter_mappings = {
        '1': '03-March', '2': '06-June', '3': '09-September', '4': '12-December',
        'Q1': '03-March', 'Q2': '06-June', 'Q3': '09-September', 'Q4': '12-December',
    }

    text = text.replace("'", "").replace(",", " ").replace("-", " ").replace("/", " ").replace(".", " ")
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()

    patterns = [
        r'(?i)\b(?:on|as of|for)?\s*(\d{1,2})?\s*(January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\s*(\d{4}|\d{2})?\b',
        r'\b(\d{4})\s+(\d{1,2})\b',
        r'\b(\d{1,2})\s+(\d{4})\b',
        r'\b(\d{4})(\d{2})\b',
        r'\b(\d{2})(\d{4})\b',
        r'\b(Q[1-4]|[1-4]Q)[\s]*(\d{2,4})\b',
        r'\b(\d{1,2})\s+(\d{1,2})\s+(\d{2,4})\b',
        r'\b(\d{2,4})\s+(\d{1,2})\s+(\d{1,2})\b',
    ]

    def try_parsing_with_formats(date_str, formats):
        for fmt in formats:
            try:
                parsed_date = datetime.datetime.strptime(date_str, fmt)
                month_num = parsed_date.strftime('%m')
                month_name = parsed_date.strftime('%B')
                year = parsed_date.strftime('%Y')
                return year, f"{month_num}-{month_name}"
            except ValueError:
                continue
        return None, None

    for pattern in patterns:
        matches = re.findall(pattern, text)
        if not matches:
            continue

        for match in matches:
            if len(match) == 3 and re.match(
                r'(?i)^(January|February|March|April|May|June|July|August|September|October|November|December|'
                r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)$',
                match[1]
            ):
                day, month_str, year = match
                day = day.strip() if day else '1'
                year = year.strip() if year else default_year
                if year:
                    if len(year) == 2:
                        year = '20' + year
                    date_str = f"{day} {month_str} {year}"
                    year_parsed, month_parsed = try_parsing_with_formats(date_str, ['%d %B %Y', '%d %b %Y'])
                    if year_parsed and month_parsed:
                        return year_parsed, month_parsed

            elif len(match) == 2 and all(part.isdigit() for part in match):
                part1, part2 = match
                if len(part1) == 4:
                    date_str = f"{part1} {part2}"
                    year_parsed, month_parsed = try_parsing_with_formats(date_str, ['%Y %m'])
                    if year_parsed and month_parsed:
                        return year_parsed, month_parsed
                elif len(part2) == 4:
                    date_str = f"{part1} {part2}"
                    year_parsed, month_parsed = try_parsing_with_formats(date_str, ['%m %Y'])
                    if year_parsed and month_parsed:
                        return year_parsed, month_parsed

            elif len(match) == 2 and all(part.isdigit() for part in match) and (len(match[0]) == 4 or len(match[1]) == 4):
                part1, part2 = match
                if len(part1) == 4 and len(part2) == 2:
                    date_str = f"{part1}{part2}"
                    try:
                        parsed_date = datetime.datetime.strptime(date_str, '%Y%m')
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        pass
                if len(part1) == 2 and len(part2) == 4:
                    date_str = f"{part2}{part1}"
                    try:
                        parsed_date = datetime.datetime.strptime(date_str, '%Y%m')
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        pass

            if len(match) == 2 and any('Q' in m for m in match):
                quarter_str, year = match
                quarter = re.sub(r'[^1-4]', '', quarter_str)
                if not quarter:
                    continue
                year = year.strip()
                if len(year) == 2:
                    year = '20' + year
                if year and quarter in quarter_mappings:
                    return year, quarter_mappings[quarter]

            if len(match) == 3 and all(part.isdigit() for part in match):
                part1, part2, part3 = match
                candidates = [
                    (f"{part1}-{part2}-{part3}", ['%d-%m-%Y', '%d-%m-%y', '%Y-%m-%d', '%y-%m-%d']),
                    (f"{part1}/{part2}/{part3}", ['%d/%m/%Y', '%d/%m/%y', '%Y/%m/%d', '%y/%m/%d']),
                    (f"{part1} {part2} {part3}", ['%d %m %Y', '%d %m %y', '%Y %m %d', '%y %m %d', '%m %d %Y', '%m %d %y']),
                ]
                for date_str, fmt_list in candidates:
                    for fmt in fmt_list:
                        try:
                            parsed_date = datetime.datetime.strptime(date_str, fmt)
                            month_num = parsed_date.strftime('%m')
                            month_name = parsed_date.strftime('%B')
                            year_val = parsed_date.strftime('%Y')
                            return year_val, f"{month_num}-{month_name}"
                        except ValueError:
                            continue

    # Fallback: If no patterns matched, try to find a month name and a 4-digit year manually.
    month_names = {
        "january": "01-January",
        "february": "02-February",
        "march": "03-March",
        "april": "04-April",
        "may": "05-May",
        "june": "06-June",
        "july": "07-July",
        "august": "08-August",
        "september": "09-September",
        "october": "10-October",
        "november": "11-November",
        "december": "12-December"
    }
    found_month = None
    for key, value in month_names.items():
        if key in text.lower():
            found_month = value
            break
    year_match = re.search(r'\b(\d{4})\b', text)
    if found_month and year_match:
        return year_match.group(1), found_month

    return default_year, None

def find_save_path(sender, subject, sender_path_table):
    rows = sender_path_table[sender_path_table['sender'].str.lower() == sender.lower()]

    if rows.empty:
        return None, None, False

    if len(rows) > 1:
        for _, row in rows.iterrows():
            coper_name = str(row.get('coper_name', '')).strip().lower()
            if coper_name and coper_name in subject.lower():
                keywords = str(row.get('keywords', '')).split(';')
                for keyword in keywords:
                    if keyword.lower() in subject.lower():
                        keyword_path = row.get('keyword_path', '')
                        return keyword_path, False, True
                save_path = row.get('save_path', '')
                return save_path, False, True
        return None, None, False
    else:
        row = rows.iloc[0]
        keywords = str(row.get('keywords', '')).split(';')
        for keyword in keywords:
            if keyword.lower() in subject.lower():
                keyword_path = row.get('keyword_path', '')
                return keyword_path, False, True
        special_case_value = str(row.get('special_case', '')).strip().lower() == 'yes'
        save_path = row.get('save_path', '')
        return save_path, special_case_value, True

def update_excel_summary(date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails):
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Summary'
        sheet.append(['Date', 'Total Emails', 'Saved in Default', 'Saved in Actual Paths', 'Not Saved'])

    sheet = workbook.active
    sheet.append([date_str, total_emails, saved_default, saved_actual, not_saved])

    if 'Failed Emails' not in workbook.sheetnames:
        failed_sheet = workbook.create_sheet('Failed Emails')
        failed_sheet.append(['Date', 'Email Address', 'Subject'])
    else:
        failed_sheet = workbook['Failed Emails']

    for email in failed_emails:
        failed_sheet.append([date_str, email['email_address'], email['subject']])

    workbook.save(EXCEL_FILE_PATH)

def save_email(item, save_path, special_case):
    try:
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        
        valid_extensions = ('.xlsx', '.xls', '.csv', '.pdf', '.doc', '.docx')
        if special_case and item.Attachments.Count > 0:
            for attachment in item.Attachments:
                if attachment.FileName.lower().endswith(valid_extensions):
                    filename_base = sanitize_filename(os.path.splitext(attachment.FileName)[0])
                    break
            else:
                filename_base = sanitize_filename(item.Subject)
        else:
            filename_base = sanitize_filename(item.Subject)
        
        extension = ".msg"
        max_filename_length = 255 - len(save_path) - len(extension) - 1
        if len(filename_base) > max_filename_length:
            filename_base = filename_base[:max_filename_length]
        
        filename = f"{filename_base}{extension}"
        full_path = os.path.join(save_path, filename)
        
        counter = 1
        while os.path.exists(full_path):
            filename = f"{filename_base}_{counter}{extension}"
            full_path = os.path.join(save_path, filename)
            counter += 1
        
        item.SaveAs(full_path, 3)
        return filename
    except pythoncom.com_error as com_err:
        error_message = f"COM Error saving email '{item.Subject}' to '{save_path}': {str(com_err)}"
        print(error_message)
        raise
    except Exception as e:
        error_message = f"General Error saving email '{item.Subject}' to '{save_path}': {str(e)}"
        print(error_message)
        raise

def process_email(item, sender_path_table, default_year, specific_date_str):
    logs = []
    failed_emails = []
    retries = 3
    processed = False

    # Attempt to extract sender email safely
    try:
        if hasattr(item, 'SenderEmailAddress') and item.SenderEmailAddress:
            sender_email = item.SenderEmailAddress.lower()
        elif hasattr(item, 'Sender') and item.Sender and hasattr(item.Sender, 'Address') and item.Sender.Address:
            sender_email = item.Sender.Address.lower()
        else:
            logs.append(f"Skipped email '{item.Subject}' due to missing sender information.")
            return logs, failed_emails
    except Exception:
        logs.append(f"Skipped email '{item.Subject}' due to error fetching sender info.")
        return logs, failed_emails

    while retries > 0 and not processed:
        try:
            year, month = extract_date_from_text(item.Subject, default_year)
            if not year or not month:
                for attachment in item.Attachments:
                    year, month = extract_date_from_text(attachment.FileName, default_year)
                    if year and month:
                        break
            year = year or default_year

            base_path, special_case, is_csv_path = find_save_path(sender_email, item.Subject, sender_path_table)
            if base_path is None:
                base_path = DEFAULT_SAVE_PATH

            # If it's a default path email
            if not is_csv_path:
                save_path = os.path.join(base_path, specific_date_str)
            else:
                if special_case:
                    save_path = os.path.join(base_path, str(year))
                else:
                    if month:
                        save_path = os.path.join(base_path, str(year), month)
                    else:
                        save_path = os.path.join(base_path, str(year))

            print(f"Email from: {sender_email}")
            print(f"Subject: {item.Subject}")
            print(f"Special Case: {special_case}")
            print(f"Save Path: {save_path}")

            filename = save_email(item, save_path, special_case)
            logs.append(f"Saved: {filename} to {save_path}")
            processed = True
        except pythoncom.com_error as com_err:
            retries -= 1
            logs.append(f"COM Error handling email '{item.Subject}' from '{sender_email}' (Code: {com_err.args})")
            if retries == 0:
                logs.append(f"Failed to save the email '{item.Subject}' from '{sender_email}' after 3 retries")
                failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
        except Exception as e:
            retries = 0
            logs.append(f"Error handling email '{item.Subject}' from '{sender_email}': {str(e)}")
            failed_emails.append({'email_address': sender_email, 'subject': item.Subject})

    return logs, failed_emails

def save_emails_from_senders_on_date(email_address, specific_date_str, sender_path_table, default_year):
    logs = []
    pythoncom.CoInitialize()
    specific_date = datetime.datetime.strptime(specific_date_str, '%Y-%m-%d').date()
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

    def find_folder_by_name(parent_folder, target_name):
        for f in parent_folder.Folders:
            if f.Name.lower() == target_name.lower():
                return f
            sub_result = find_folder_by_name(f, target_name)
            if sub_result is not None:
                return sub_result
        return None

    # Hard-code the folder name to "NAV and Performance"
    folder_to_find = "NAV and Performance"

    inbox = None
    target_folder = None

    for store in outlook.Stores:
        if store.DisplayName.lower() == email_address.lower() or store.ExchangeStoreType == 3:
            try:
                root_folder = store.GetRootFolder()
                inbox = next((folder for folder in root_folder.Folders if folder.Name.lower() == "inbox"), None)
                if inbox and folder_to_find:
                    target_folder = find_folder_by_name(inbox, folder_to_find)
                break
            except AttributeError as e:
                logs.append(f"Error accessing folders: {str(e)}")
                continue

    if not inbox:
        logs.append(f"No Inbox found for the account with email address: {email_address}")
        pythoncom.CoUninitialize()
        with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
            f.writelines("\n".join(logs))
        return
    else:
        logs.append("Inbox found successfully.")

    if folder_to_find and not target_folder:
        logs.append(f"No '{folder_to_find}' folder found as a subfolder of Inbox.")
    elif folder_to_find and target_folder:
        logs.append(f"'{folder_to_find}' folder found successfully.")

    def get_items_for_folder(folder, date):
        filtered_items = []
        if folder:
            items = folder.Items
            items.Sort("[ReceivedTime]", True)
            items = items.Restrict(
                f"[ReceivedTime] >= '{date.strftime('%m/%d/%Y')} 00:00 AM' AND [ReceivedTime] <= '{date.strftime('%m/%d/%Y')} 11:59 PM'"
            )
            for item in items:
                filtered_items.append(item)
        return filtered_items

    inbox_items = get_items_for_folder(inbox, specific_date) if inbox else []
    target_items = get_items_for_folder(target_folder, specific_date) if target_folder else []

    all_items = inbox_items + target_items
    logs.append(f"Total emails found: {len(all_items)} (Inbox: {len(inbox_items)}, '{folder_to_find}': {len(target_items)})")

    total_emails = len(all_items)
    saved_default, saved_actual, not_saved = 0, 0, 0
    failed_emails = []

    for item in all_items:
        email_logs, email_failed_emails = process_email(item, sender_path_table, default_year, specific_date_str)
        logs.extend(email_logs)
        failed_emails.extend(email_failed_emails)
        if any(DEFAULT_SAVE_PATH in log for log in email_logs):
            saved_default += 1
        else:
            saved_actual += 1

    pythoncom.CoUninitialize()
    with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
        f.writelines("\n".join(logs))

    update_excel_summary(specific_date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails)
    print("Process completed for", specific_date_str, ". Check logs.txt and email_summary.xlsx for details.")

# ------------------ Calendar Helper Functions ------------------ #
# These functions use tkinter and tkcalendar for a graphical date selection.
def select_date_via_calendar(title="Select Date"):
    try:
        import tkinter as tk
        from tkcalendar import Calendar
    except ImportError:
        print("tkinter and/or tkcalendar not available. Falling back to text input.")
        return None

    selected_date = []

    def on_select():
        selected_date.append(cal.get_date())
        root.destroy()

    root = tk.Tk()
    root.title(title)
    cal = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal.pack(padx=10, pady=10)
    btn = tk.Button(root, text="Select", command=on_select)
    btn.pack(pady=10)
    root.mainloop()
    if selected_date:
        return selected_date[0]
    return None

# ------------------ Main Interactive Section ------------------ #
if __name__ == '__main__':
    # Hard-code the email account here:
    email_address = 'your_email@domain.com'

    # Check if calendar selection is available
    try:
        import tkinter as tk
        from tkcalendar import Calendar
        calendar_available = True
    except ImportError:
        calendar_available = False

    use_calendar = False
    if calendar_available:
        cal_choice = input("Do you want to use the calendar for date selection? (Y/N): ").strip().lower()
        if cal_choice == 'y':
            use_calendar = True

    print("\nSelect date option:")
    print("1: Yesterday")
    print("2: Specific Date")
    print("3: Date Range")
    option = input("Enter option number: ").strip()

    date_list = []

    if option == '1':
        yesterday = datetime.date.today() - datetime.timedelta(days=1)
        date_list = [yesterday.strftime('%Y-%m-%d')]
    elif option == '2':
        if use_calendar:
            selected = select_date_via_calendar("Select a Date")
            if not selected:
                print("No date selected. Exiting.")
                exit(1)
            date_list = [selected]
        else:
            date_input = input("Enter the date (YYYY-MM-DD): ").strip()
            try:
                datetime.datetime.strptime(date_input, '%Y-%m-%d')
                date_list = [date_input]
            except ValueError:
                print("Invalid date format. Please use YYYY-MM-DD.")
                exit(1)
    elif option == '3':
        if use_calendar:
            start = select_date_via_calendar("Select Start Date")
            end = select_date_via_calendar("Select End Date")
        else:
            start = input("Enter the start date (YYYY-MM-DD): ").strip()
            end = input("Enter the end date (YYYY-MM-DD): ").strip()
        try:
            start_date = datetime.datetime.strptime(start, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end, '%Y-%m-%d').date()
        except ValueError:
            print("Invalid date format. Please use YYYY-MM-DD.")
            exit(1)
        if start_date > end_date:
            print("Start date must not be after end date.")
            exit(1)
        # Generate list of dates in the range
        delta = end_date - start_date
        date_list = [(start_date + datetime.timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]
    else:
        print("Invalid option selected.")
        exit(1)

    default_year = input("Enter the default year (YYYY): ").strip()
    if not (default_year.isdigit() and len(default_year) == 4):
        print("Invalid year format. Please enter a year in YYYY format.")
        exit(1)

    # Load sender_path_table
    csv_file_path = SENDER_PATH_TABLE_PATH
    if not os.path.exists(csv_file_path):
        print("CSV file not found. Please provide a valid file path.")
        exit(1)
    try:
        sender_path_table = pd.read_csv(csv_file_path, encoding='utf-8')
    except UnicodeDecodeError:
        try:
            sender_path_table = pd.read_csv(csv_file_path, encoding='latin1')
        except Exception as e:
            print(f"Error reading the CSV file: {e}")
            exit(1)
    sender_path_table.columns = sender_path_table.columns.str.lower()

    os.makedirs(DEFAULT_SAVE_PATH, exist_ok=True)

    # Process each selected date
    for d in date_list:
        print("\nProcessing emails for:", d)
        save_emails_from_senders_on_date(email_address, d, sender_path_table, default_year)
