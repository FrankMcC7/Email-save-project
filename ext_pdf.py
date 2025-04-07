import os
import re
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import glob
from datetime import datetime
import traceback

def extract_fund_name(pdf_path):
    """
    Extract the fund name from the third line of the first page.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) > 0:
                first_page_text = pdf.pages[0].extract_text()
                if first_page_text:
                    # Split text by newlines and get the third line if available
                    lines = first_page_text.split('\n')
                    
                    # First try getting the third line if it's not empty
                    if len(lines) >= 3 and lines[2].strip():
                        return lines[2].strip()
                    
                    # Try looking for fund name patterns in the first 10 lines
                    for i in range(min(10, len(lines))):
                        line = lines[i].strip()
                        # Skip empty lines or very short lines
                        if len(line) < 3:
                            continue
                            
                        # Check for fund name indicators
                        if any(indicator in line.lower() for indicator in ['fund', 'sicav', 's.c.a', 'l.p.', 'partners group']):
                            return line
        
        # Fallback to extracting text from first few pages and looking for fund name patterns
        all_text = extract_text_from_pdf(pdf_path, max_pages=3)
        
        # Try to find title lines that might contain fund names
        lines = all_text.split('\n')
        for line in lines[:20]:  # Look only in first 20 lines
            line = line.strip()
            if len(line) > 10 and any(indicator in line.lower() for indicator in ['fund', 'sicav', 's.c.a', 'l.p.', 'partners group']):
                return line
        
        # If all else fails, try to match common fund name patterns
        fund_patterns = [
            r"([A-Za-z0-9\s\-\.&]+(?:S\.C\.A\.|SICAV|Fund|L\.P\.))",
            r"([A-Za-z]+\s+[A-Za-z]+\s+[A-Za-z]+(?:\s+[IVX]+)?)"
        ]
        
        for pattern in fund_patterns:
            match = re.search(pattern, all_text)
            if match:
                return match.group(1).strip()
    except Exception as e:
        print(f"Error extracting fund name: {str(e)}")
    
    return "Unknown Fund"

def extract_text_from_pdf(pdf_path, max_pages=None, start_page=0):
    """
    Extract text from PDF, optionally limiting to max_pages starting from start_page.
    """
    try:
        all_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            # Determine which pages to extract
            if max_pages is None:
                pages_to_extract = pdf.pages[start_page:]
            else:
                end_page = min(start_page + max_pages, len(pdf.pages))
                pages_to_extract = pdf.pages[start_page:end_page]
                
            for page in pages_to_extract:
                page_text = page.extract_text() or ""
                all_text += page_text + "\n"
        return all_text
    except Exception as e:
        print(f"Error extracting text from PDF: {str(e)}")
        return ""

def direct_table_extraction(pdf_path):
    """
    Directly target the Key Figures table and extract Net asset value.
    This approach focuses specifically on the exact table structure seen in examples.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check pages 1-5 for the key figures table
            for page_num in range(min(5, len(pdf.pages))):
                page = pdf.pages[page_num]
                
                # Get the text to check for Key Figures section
                page_text = page.extract_text() or ""
                
                # Only process pages with "Key figures" or "key figures"
                if "Key figures" in page_text or "key figures" in page_text.lower():
                    # Extract all tables from this page
                    tables = page.extract_tables()
                    
                    # Process each table
                    for table in tables:
                        if not table or len(table) < 3:  # Skip small tables
                            continue
                        
                        # Look for NAV row in this table
                        for row_idx, row in enumerate(table):
                            # Skip rows with fewer than 2 columns
                            if len(row) < 3:
                                continue
                                
                            # Convert to string and check if this is the NAV row
                            row_str = [str(cell).strip() if cell is not None else "" for cell in row]
                            first_cell = row_str[0].lower()
                            
                            if "net asset value" in first_cell or "net asset" in first_cell:
                                # Found the NAV row - now get header and values
                                
                                # Look for header row (containing dates or periods)
                                header_row = None
                                for i in range(row_idx):
                                    potential_header = table[i]
                                    if len(potential_header) >= len(row):
                                        header_text = " ".join([str(cell).strip() if cell is not None else "" for cell in potential_header])
                                        # Check if it contains dates or period indicators
                                        if (re.search(r'\b\d{1,2}[\.\/\s]+(?:[A-Za-z]+|\d{1,2})[\.\/\s]+\d{4}\b', header_text) or
                                            re.search(r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b', header_text, re.IGNORECASE) or
                                            re.search(r'\b(?:Q[1-4]|20\d\d)\b', header_text)):
                                            header_row = potential_header
                                            break
                                
                                # If no proper header found, use the row above if available
                                if header_row is None and row_idx > 0:
                                    header_row = table[row_idx - 1]
                                
                                # Get values from NAV row (skip first column which is the label)
                                values = []
                                for cell_idx in range(1, len(row)):
                                    if row[cell_idx] is not None:
                                        # Clean and convert the value
                                        try:
                                            value_str = str(row[cell_idx]).strip()
                                            value = clean_number(value_str)
                                            if value is not None:
                                                values.append((cell_idx, value))
                                        except:
                                            pass
                                
                                # Need at least two values
                                if len(values) >= 2:
                                    # Get corresponding headers
                                    idx1, value1 = values[0]
                                    idx2, value2 = values[1]
                                    
                                    if header_row:
                                        header1 = str(header_row[idx1]).strip() if idx1 < len(header_row) and header_row[idx1] is not None else "Period 1"
                                        header2 = str(header_row[idx2]).strip() if idx2 < len(header_row) and header_row[idx2] is not None else "Period 2"
                                    else:
                                        # Extract dates from page text if headers not found
                                        date_matches = re.findall(r'\b\d{1,2}[\.\/\s]+(?:[A-Za-z]+|\d{1,2})[\.\/\s]+\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b|\b(?:Q[1-4]|20\d\d)\b', page_text)
                                        if len(date_matches) >= 2:
                                            header1 = date_matches[0]
                                            header2 = date_matches[1]
                                        else:
                                            header1 = "Period 1"
                                            header2 = "Period 2"
                                    
                                    return header1, value1, header2, value2
    except Exception as e:
        print(f"Error in direct table extraction: {str(e)}")
        
    return None, None, None, None

def image_based_extraction(pdf_path):
    """
    Attempt to extract data focusing on visual clues like row with "Net asset value".
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check the first few pages
            for page_num in range(min(5, len(pdf.pages))):
                page = pdf.pages[page_num]
                page_text = page.extract_text() or ""
                
                # Only process if it might contain Key Figures
                if "key" in page_text.lower() and "figure" in page_text.lower():
                    
                    # Extract lines that might contain dates and NAV values
                    lines = page_text.split('\n')
                    date_line = None
                    nav_line = None
                    
                    for line in lines:
                        # Check if line contains dates
                        if re.search(r'\b\d{1,2}[\.\/\s]+(?:[A-Za-z]+|\d{1,2})[\.\/\s]+\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', line):
                            date_line = line
                        
                        # Check if line contains Net asset value
                        if "net asset value" in line.lower():
                            nav_line = line
                    
                    # If we found both date and NAV lines
                    if date_line and nav_line:
                        # Extract dates
                        dates = re.findall(r'\b\d{1,2}[\.\/\s]+(?:[A-Za-z]+|\d{1,2})[\.\/\s]+\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', date_line)
                        if len(dates) < 2:
                            # Try adding quarter or year patterns
                            additional_dates = re.findall(r'\b(?:Q[1-4]|20\d\d)\b', date_line)
                            dates.extend(additional_dates)
                        
                        # Extract numbers from NAV line
                        numbers = re.findall(r'\b\d[\d\s\.\',]+\b', nav_line)
                        values = []
                        
                        for num in numbers:
                            value = clean_number(num)
                            if value is not None and value > 10000:  # NAV values are typically large
                                values.append(value)
                        
                        if len(dates) >= 2 and len(values) >= 2:
                            return dates[0], values[0], dates[1], values[1]
    except Exception as e:
        print(f"Error in image-based extraction: {str(e)}")
        
    return None, None, None, None

def scan_text_for_nav(pdf_path):
    """
    Scan the PDF text for Net Asset Value mentions and extract data.
    This is a fallback approach when table extraction fails.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Extract text from the first few pages
            all_text = ""
            for page_num in range(min(5, len(pdf.pages))):
                page_text = pdf.pages[page_num].extract_text() or ""
                all_text += page_text + "\n"
                
                # Check each page individually first
                lines = page_text.split('\n')
                
                # 1. First, look for "Net asset value" line specifically
                for i, line in enumerate(lines):
                    if "net asset value" in line.lower():
                        # Found NAV line - extract numbers
                        number_matches = re.findall(r"[\d',\.]+", line)
                        values = []
                        
                        for num in number_matches:
                            value = clean_number(num)
                            if value is not None and value > 1000:
                                values.append(value)
                        
                        if len(values) >= 2:
                            # Look for date headers in previous lines
                            date_matches = []
                            for j in range(max(0, i-5), i):
                                date_match = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', lines[j])
                                if date_match:
                                    date_matches.extend(date_match)
                            
                            if len(date_matches) >= 2:
                                return date_matches[0], values[0], date_matches[1], values[1]
                            else:
                                # Look for dates in the entire page
                                date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', page_text)
                                if len(date_matches) >= 2:
                                    return date_matches[0], values[0], date_matches[1], values[1]
                                else:
                                    return "Period 1", values[0], "Period 2", values[1]
            
            # If page-by-page approach failed, try with whole text
            nav_matches = re.finditer(r'(?:[Nn]et\s+[Aa]sset\s+[Vv]alue|NAV)', all_text)
            
            for match in nav_matches:
                # Extract text around the match
                start_pos = max(0, match.start() - 100)
                end_pos = min(len(all_text), match.end() + 300)
                context = all_text[start_pos:end_pos]
                
                # Find dates in the context
                dates = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', context)
                
                # Find large numbers that could be NAV values
                number_matches = re.findall(r"[\d',\.]+", context)
                values = []
                
                for num in number_matches:
                    value = clean_number(num)
                    if value is not None and value > 10000:  # NAV values are typically large
                        values.append(value)
                
                if len(dates) >= 2 and len(values) >= 2:
                    return dates[0], values[0], dates[1], values[1]
    except Exception as e:
        print(f"Error in text scanning: {str(e)}")
        
    return None, None, None, None

def advanced_table_analysis(pdf_path):
    """
    Use more advanced table parsing to handle complex table structures.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check pages 1-5 for the key figures table
            for page_num in range(min(5, len(pdf.pages))):
                page = pdf.pages[page_num]
                page_text = page.extract_text() or ""
                
                # Look for any mention of "Net asset value" or similar in the page
                if "net asset value" in page_text.lower() or "key figures" in page_text.lower():
                    
                    # Try different table extraction settings
                    for table_settings in [
                        {'vertical_strategy': 'lines', 'horizontal_strategy': 'lines'},
                        {'vertical_strategy': 'text', 'horizontal_strategy': 'text'},
                        {'vertical_strategy': 'lines', 'horizontal_strategy': 'text'},
                        {'vertical_strategy': 'text', 'horizontal_strategy': 'lines'}
                    ]:
                        try:
                            tables = page.extract_tables(table_settings)
                            
                            for table in tables:
                                if not table or len(table) < 3:
                                    continue
                                    
                                # Find rows with Net asset value
                                for row_idx, row in enumerate(table):
                                    if len(row) < 3:
                                        continue
                                        
                                    row_str = [str(cell).strip() if cell is not None else "" for cell in row]
                                    row_text = " ".join(row_str).lower()
                                    first_cell = row_str[0].lower() if row_str else ""
                                    
                                    # Look specifically for "Net asset value" in first column
                                    if "net asset value" in first_cell:
                                        # Get values from this row
                                        values = []
                                        for cell_idx, cell in enumerate(row):
                                            if cell_idx > 0 and cell is not None:  # Skip first column (label)
                                                value = clean_number(str(cell))
                                                if value is not None and value > 1000:  # NAV is typically large
                                                    values.append((cell_idx, value))
                                        
                                        if len(values) >= 2:
                                            # Find header row
                                            header_row = None
                                            for i in range(row_idx):
                                                potential_header = table[i]
                                                if len(potential_header) >= len(row):
                                                    header_text = " ".join([str(cell).strip() if cell is not None else "" for cell in potential_header])
                                                    # Check for date patterns
                                                    if re.search(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', header_text):
                                                        header_row = potential_header
                                                        break
                                            
                                            # Get header labels
                                            idx1, value1 = values[0]
                                            idx2, value2 = values[1]
                                            
                                            if header_row:
                                                header1 = str(header_row[idx1]).strip() if idx1 < len(header_row) and header_row[idx1] is not None else "Period 1"
                                                header2 = str(header_row[idx2]).strip() if idx2 < len(header_row) and header_row[idx2] is not None else "Period 2"
                                            else:
                                                # Extract dates from page text if headers not found
                                                date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', page_text)
                                                if len(date_matches) >= 2:
                                                    header1 = date_matches[0]
                                                    header2 = date_matches[1]
                                                else:
                                                    header1 = "Period 1"
                                                    header2 = "Period 2"
                                            
                                            # Ensure values are reasonable for NAV figures
                                            if value1 > 1000 and value2 > 1000:
                                                return header1, value1, header2, value2
                        except:
                            continue  # Try next table settings if this one fails
                            
                    # Handle the case where "Net asset value" is not a separate row but part of a column
                    # This is specifically for the alternative table format shown in the screenshots
                    try:
                        # Find dates in page text for header information
                        date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', page_text)
                        
                        # Find the line with "Net asset value" and extract numbers
                        net_asset_lines = []
                        for line in page_text.split('\n'):
                            if "net asset value" in line.lower():
                                net_asset_lines.append(line)
                        
                        for line in net_asset_lines:
                            # Extract numbers from this line
                            number_matches = re.findall(r"[\d',\.]+", line)
                            values = []
                            
                            for num in number_matches:
                                value = clean_number(num)
                                if value is not None and value > 1000:
                                    values.append(value)
                            
                            if len(values) >= 2 and len(date_matches) >= 2:
                                return date_matches[0], values[0], date_matches[1], values[1]
                    except:
                        pass
    except Exception as e:
        print(f"Error in advanced table analysis: {str(e)}")
        
    return None, None, None, None        
                                        return header1, value1, header2, value2
                    except:
                        continue  # Try next table settings if this one fails
    except Exception as e:
        print(f"Error in advanced table analysis: {str(e)}")
        
    return None, None, None, None

def last_resort_extraction(pdf_path):
    """
    Absolute last resort approach - specifically look for patterns like in the Excel screenshot.
    This approach focuses on known formats from the examples provided.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check first 3 pages
            for page_num in range(min(3, len(pdf.pages))):
                page = pdf.pages[page_num]
                page_text = page.extract_text() or ""
                
                # Look specifically for "Net asset value" in a line
                lines = page_text.split('\n')
                
                # First approach: Look for the exact line format with Net asset value
                for line in lines:
                    if "net asset value" in line.lower():
                        # Try to extract exactly two numbers from this line
                        number_matches = re.findall(r"[\d',\.]+", line)
                        values = []
                        
                        for num in number_matches:
                            value = clean_number(num)
                            if value is not None and value > 10000:  # NAV values are typically large
                                values.append(value)
                        
                        if len(values) >= 2:
                            # Extract dates from the page text
                            date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', page_text)
                            
                            if len(date_matches) >= 2:
                                return date_matches[0], values[0], date_matches[1], values[1]
                            else:
                                # If no dates found, just use Period 1 and Period 2
                                return "Period 1", values[0], "Period 2", values[1]
                
                # Second approach: Look for numeric patterns in lines with dates
                date_line_idx = -1
                for i, line in enumerate(lines):
                    if re.search(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', line):
                        date_line_idx = i
                        break
                
                if date_line_idx >= 0:
                    # Found a line with dates, now look for "Net asset value" within the next 15 lines
                    for i in range(date_line_idx + 1, min(date_line_idx + 15, len(lines))):
                        if "net asset value" in lines[i].lower():
                            # Found the NAV line
                            number_matches = re.findall(r"[\d',\.]+", lines[i])
                            values = []
                            
                            for num in number_matches:
                                value = clean_number(num)
                                if value is not None and value > 10000:
                                    values.append(value)
                            
                            if len(values) >= 2:
                                # Extract dates from the date line
                                date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', lines[date_line_idx])
                                
                                if len(date_matches) >= 2:
                                    return date_matches[0], values[0], date_matches[1], values[1]
            
            # Third approach: Process all lines looking for specific formats like in the Excel
            for page_num in range(min(5, len(pdf.pages))):
                page_text = pdf.pages[page_num].extract_text() or ""
                lines = page_text.split('\n')
                
                for line in lines:
                    # Look for lines that match the format "some text number1 number2"
                    if "asset" in line.lower() or "value" in line.lower():
                        number_matches = re.findall(r"[\d',\.]+", line)
                        values = []
                        
                        for num in number_matches:
                            value = clean_number(num)
                            if value is not None and value > 10000:
                                values.append(value)
                        
                        if len(values) >= 2:
                            # Extract dates from the page text
                            date_matches = re.findall(r'\b\d{1,2}[\.\/]\d{1,2}[\.\/]\d{4}\b|\b\d{2}\.\d{2}\.\d{4}\b', page_text)
                            
                            if len(date_matches) >= 2:
                                return date_matches[0], values[0], date_matches[1], values[1]
    except Exception as e:
        print(f"Error in last resort extraction: {str(e)}")
        
    return None, None, None, None

def clean_number(value_str):
    """
    Clean and convert a string value to a float.
    Handles various number formats including apostrophes, commas, spaces.
    """
    try:
        if value_str is None:
            return None
            
        # If it's already a number, return it
        if isinstance(value_str, (int, float)):
            return float(value_str)
            
        if isinstance(value_str, str):
            # Handle percentage values - skip these
            if "%" in value_str:
                return None
            
            # Skip non-numeric text
            if len(value_str.strip()) == 0 or value_str.strip().isalpha():
                return None
                
            # Remove common separators and non-numeric characters
            cleaned = value_str.replace("'", "").replace(",", "").replace(" ", "").replace("x", "")
            
            # Handle cases where there's no digits
            if not any(c.isdigit() for c in cleaned):
                return None
                
            # Handle any remaining non-numeric characters (except decimal point)
            final_value = ""
            decimal_found = False
            for char in cleaned:
                if char.isdigit():
                    final_value += char
                elif char == "." and not decimal_found:
                    final_value += char
                    decimal_found = True
                    
            if final_value:
                # Convert to float with additional validation
                try:
                    value = float(final_value)
                    
                    # If we have a non-zero value and it's reasonably large (NAV values are typically large)
                    # This helps avoid misidentifying small numbers as NAV values
                    if value > 100:  # NAV is typically in thousands or millions
                        return value
                except:
                    pass
    except:
        pass
        
    return None

def process_pdf(pdf_path):
    """
    Process a single PDF file to extract fund name and NAV values.
    Uses multiple fallback approaches to maximize success.
    """
    try:
        # Extract fund name
        fund_name = extract_fund_name(pdf_path)
        
        # Try multiple extraction approaches in sequence
        period1_label = period2_label = None
        period1_nav = period2_nav = None
        
        # First approach: Direct table extraction targeting Key Figures table
        period1_label, period1_nav, period2_label, period2_nav = direct_table_extraction(pdf_path)
        
        # Check if values seem reasonable based on fund name
        if period1_nav is not None and period2_nav is not None:
            if period1_nav < 1000 and period2_nav < 1000 and "GLOBAL VALUE" not in fund_name and "MERCATI" not in fund_name:
                # Values seem too small for a typical fund NAV, might be incorrect
                period1_nav = period2_nav = None
        
        # Second approach: Advanced table analysis with different settings
        if period1_nav is None or period2_nav is None:
            period1_label, period1_nav, period2_label, period2_nav = advanced_table_analysis(pdf_path)
            
        # Third approach: Text scanning for NAV mentions
        if period1_nav is None or period2_nav is None:
            period1_label, period1_nav, period2_label, period2_nav = scan_text_for_nav(pdf_path)
        
        # Fourth approach: Image-based extraction if first approach failed
        if period1_nav is None or period2_nav is None:
            period1_label, period1_nav, period2_label, period2_nav = image_based_extraction(pdf_path)
        
        # Last resort: Desperate pattern matching
        if period1_nav is None or period2_nav is None:
            period1_label, period1_nav, period2_label, period2_nav = last_resort_extraction(pdf_path)
        
        # Make sure date formats are consistent if we have dates
        if period1_label and period2_label:
            # If dates are in DD.MM.YYYY format, standardize them
            if re.match(r'\d{2}\.\d{2}\.\d{4}', period1_label):
                period1_label = period1_label.replace('.', '/')
            if re.match(r'\d{2}\.\d{2}\.\d{4}', period2_label):
                period2_label = period2_label.replace('.', '/')
                
            # If we have Periods instead of dates, try to find real dates
            if period1_label == "Period 1" or period2_label == "Period 2":
                with pdfplumber.open(pdf_path) as pdf:
                    # Check first few pages for dates
                    for page_num in range(min(3, len(pdf.pages))):
                        page_text = pdf.pages[page_num].extract_text() or ""
                        date_matches = re.findall(r'\b\d{2}\.\d{2}\.\d{4}\b|\b\d{2}/\d{2}/\d{4}\b', page_text)
                        if len(date_matches) >= 2:
                            if period1_label == "Period 1":
                                period1_label = date_matches[0].replace('.', '/')
                            if period2_label == "Period 2":
                                period2_label = date_matches[1].replace('.', '/')
                            break
        
        return {
            'Fund Name': fund_name,
            'Period 1 Label': period1_label if period1_label else "Period 1",
            'Period 1 NAV': period1_nav,
            'Period 2 Label': period2_label if period2_label else "Period 2",
            'Period 2 NAV': period2_nav,
            'PDF Filename': os.path.basename(pdf_path)
        }
    except Exception as e:
        print(f"Error processing {pdf_path}: {str(e)}")
        traceback.print_exc()
        return {
            'Fund Name': fund_name if 'fund_name' in locals() else f"Error: {os.path.basename(pdf_path)}",
            'Period 1 Label': 'Period 1',
            'Period 1 NAV': None,
            'Period 2 Label': 'Period 2',
            'Period 2 NAV': None,
            'PDF Filename': os.path.basename(pdf_path)
        }

def format_excel(excel_path):
    """
    Apply formatting to the Excel file for better readability.
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Apply header formatting
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format numeric values
        for row in ws.iter_rows(min_row=2):
            for cell in row[2:]:  # Apply only to NAV columns
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
        
        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"Error formatting Excel file: {str(e)}")
        return False

def main():
    # Folder containing PDF files
    pdf_folder = input("Enter the path to the folder containing PDF files: ")
    
    # Check if the folder exists
    if not os.path.isdir(pdf_folder):
        print(f"Error: Folder '{pdf_folder}' does not exist.")
        return
    
    # Find all PDF files in the folder
    pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
    
    if not pdf_files:
        print(f"No PDF files found in folder '{pdf_folder}'.")
        return
    
    print(f"Found {len(pdf_files)} PDF files. Processing...")
    
    # Process each PDF file
    results = []
    success_count = 0
    failure_count = 0
    
    for i, pdf_path in enumerate(pdf_files):
        print(f"Processing {i+1}/{len(pdf_files)}: {os.path.basename(pdf_path)}...")
        result = process_pdf(pdf_path)
        results.append(result)
        
        if result['Period 1 NAV'] is not None and result['Period 2 NAV'] is not None:
            success_count += 1
        else:
            failure_count += 1
    
    # Create DataFrame from results
    df = pd.DataFrame(results)
    
    # Create output Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(pdf_folder, f"Fund_NAV_Summary_{timestamp}.xlsx")
    
    # Export to Excel
    df.to_excel(output_path, index=False)
    
    # Format the Excel file
    format_excel(output_path)
    
    print(f"\nProcessing complete. Results saved to: {output_path}")
    print(f"Success: {success_count}/{len(pdf_files)} ({success_count/len(pdf_files)*100:.1f}%)")
    print(f"Failed: {failure_count}/{len(pdf_files)} ({failure_count/len(pdf_files)*100:.1f}%)")

if __name__ == "__main__":
    main()
