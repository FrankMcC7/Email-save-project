import os
import datetime
import re
import pythoncom
import win32com.client
import pandas as pd
import json
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_socketio import SocketIO
import openpyxl
import unicodedata
from dateparser import parse as date_parse

app = Flask(__name__)
app.secret_key = 'supersecretkey'
socketio = SocketIO(app)

# Load configurations from a JSON file
try:
    with open('config.json', 'r') as f:
        config = json.load(f)
except FileNotFoundError:
    config = {}
    print("Configuration file 'config.json' not found. Using default settings.")

DEFAULT_SAVE_PATH = config.get('DEFAULT_SAVE_PATH', 'path_to_default_folder')
LOG_FILE_PATH = config.get('LOG_FILE_PATH', 'logs.txt')
EXCEL_FILE_PATH = config.get('EXCEL_FILE_PATH', 'email_summary.xlsx')

def sanitize_filename(filename):
    # Normalize unicode characters to their closest ASCII equivalent (e.g., é -> e)
    normalized_filename = unicodedata.normalize('NFKD', filename).encode('ASCII', 'ignore').decode('ASCII')

    # Replace common problematic characters with underscores or remove them
    sanitized = re.sub(r'[<>:"/\\|?*\[\]\'`~!@#$%^&*()+={};,]', '_', normalized_filename)

    # Replace dots (.) followed by a space or end of the string with an underscore, except for file extensions
    sanitized = re.sub(r'\.(?=\s|$)', '_', sanitized)

    # Replace multiple underscores with a single underscore
    sanitized = re.sub(r'_+', '_', sanitized)

    # Trim leading and trailing underscores or spaces
    sanitized = sanitized.strip(' _')

    # Limit filename length (255 chars total)
    sanitized = sanitized[:255]

    return sanitized

def extract_date_from_text(text, default_year=None):
    import datetime

    # Map quarters to months
    quarter_mappings = {
        '1': '03-March', '2': '06-June', '3': '09-September', '4': '12-December',
        'Q1': '03-March', 'Q2': '06-June', 'Q3': '09-September', 'Q4': '12-December',
    }

    # Normalize separators
    text = text.replace("'", "").replace(",", " ").replace("-", " ").replace("/", " ").replace(".", " ")
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()

    patterns = [
        r'(?i)\b(?:on|as of|for)?\s*(\d{1,2})?\s*(January|February|March|April|May|June|July|August|September|October|November|December|'
        r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\s*(\d{4}|\d{2})?\b',
        r'\b(\d{4})\s+(\d{1,2})\b',
        r'\b(\d{1,2})\s+(\d{4})\b',
        r'\b(\d{4})(\d{2})\b',
        r'\b(\d{2})(\d{4})\b',
        r'\b(Q[1-4]|[1-4]Q)[\s]*(\d{2,4})\b',
        r'\b(\d{1,2})\s+(\d{1,2})\s+(\d{2,4})\b',
        r'\b(\d{2,4})\s+(\d{1,2})\s+(\d{1,2})\b',
    ]

    def try_parsing_with_formats(date_str, formats):
        for fmt in formats:
            try:
                parsed_date = datetime.datetime.strptime(date_str, fmt)
                month_num = parsed_date.strftime('%m')
                month_name = parsed_date.strftime('%B')
                year = parsed_date.strftime('%Y')
                return year, f"{month_num}-{month_name}"
            except ValueError:
                continue
        return None, None

    # Try regex-based patterns first
    for pattern in patterns:
        matches = re.findall(pattern, text)
        if not matches:
            continue

        for match in matches:
            # Month name pattern
            if len(match) == 3 and re.match(
                r'(?i)^(January|February|March|April|May|June|July|August|September|October|November|December|'
                r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)$',
                match[1]
            ):
                day, month_str, year = match
                day = day.strip() if day else '1'
                year = year.strip() if year else default_year
                if year:
                    if len(year) == 2:
                        year = '20' + year
                    date_str = f"{day} {month_str} {year}"
                    year_parsed, month_parsed = try_parsing_with_formats(date_str, ['%d %B %Y', '%d %b %Y'])
                    if year_parsed and month_parsed:
                        return year_parsed, month_parsed

            # YYYY MM or MM YYYY patterns
            elif len(match) == 2 and all(part.isdigit() for part in match):
                part1, part2 = match
                # YYYY MM
                if len(part1) == 4:
                    date_str = f"{part1} {part2}"
                    year_parsed, month_parsed = try_parsing_with_formats(date_str, ['%Y %m'])
                    if year_parsed and month_parsed:
                        return year_parsed, month_parsed
                # MM YYYY
                elif len(part2) == 4:
                    date_str = f"{part1} {part2}"
                    year_parsed, month_parsed = try_parsing_with_formats(date_str, ['%m %Y'])
                    if year_parsed and month_parsed:
                        return year_parsed, month_parsed

            # Compact YYYYMM or MMYYYY
            elif len(match) == 2 and all(part.isdigit() for part in match) and (len(match[0]) == 4 or len(match[1]) == 4):
                part1, part2 = match
                # YYYYMM
                if len(part1) == 4 and len(part2) == 2:
                    date_str = f"{part1}{part2}"
                    try:
                        parsed_date = datetime.datetime.strptime(date_str, '%Y%m')
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        pass
                # MMYYYY
                if len(part1) == 2 and len(part2) == 4:
                    date_str = f"{part2}{part1}"
                    try:
                        parsed_date = datetime.datetime.strptime(date_str, '%Y%m')
                        month_num = parsed_date.strftime('%m')
                        month_name = parsed_date.strftime('%B')
                        year = parsed_date.strftime('%Y')
                        return year, f"{month_num}-{month_name}"
                    except ValueError:
                        pass

            # Quarter patterns
            if len(match) == 2 and any('Q' in m for m in match):
                quarter_str, year = match
                quarter = re.sub(r'[^1-4]', '', quarter_str)
                if not quarter:
                    continue
                year = year.strip()
                if len(year) == 2:
                    year = '20' + year
                if year and quarter in quarter_mappings:
                    return year, quarter_mappings[quarter]

            # Numeric date patterns (DD MM YYYY, YYYY MM DD, etc.)
            if len(match) == 3 and all(part.isdigit() for part in match):
                part1, part2, part3 = match
                candidates = [
                    (f"{part1}-{part2}-{part3}", ['%d-%m-%Y', '%d-%m-%y', '%Y-%m-%d', '%y-%m-%d']),
                    (f"{part1}/{part2}/{part3}", ['%d/%m/%Y', '%d/%m/%y', '%Y/%m/%d', '%y/%m/%d']),
                    (f"{part1} {part2} {part3}", ['%d %m %Y', '%d %m %y', '%Y %m %d', '%y %m %d', '%m %d %Y', '%m %d %y']),
                ]
                for date_str, fmt_list in candidates:
                    for fmt in fmt_list:
                        try:
                            parsed_date = datetime.datetime.strptime(date_str, fmt)
                            month_num = parsed_date.strftime('%m')
                            month_name = parsed_date.strftime('%B')
                            year_val = parsed_date.strftime('%Y')
                            return year_val, f"{month_num}-{month_name}"
                        except ValueError:
                            continue

    # Fallback to dateparser
    parsed_date = date_parse(text, settings={'REQUIRE_PARTS': ['year', 'month'], 'PREFER_DATES_FROM': 'past'})
    if parsed_date:
        year = parsed_date.strftime('%Y')
        month_num = parsed_date.strftime('%m')
        month_name = parsed_date.strftime('%B')
        if not year and default_year:
            year = default_year
        return year, f"{month_num}-{month_name}"

    # If no date found
    return default_year, None

def find_save_path(sender, subject, sender_path_table):
    # Check if the sender exists in the CSV file
    rows = sender_path_table[sender_path_table['sender'].str.lower() == sender.lower()]

    # If the sender is not found in the CSV, treat it as a default path email
    if rows.empty:
        return None, None, False

    # If multiple entries, apply coper_name logic
    if len(rows) > 1:
        for _, row in rows.iterrows():
            coper_name = str(row.get('coper_name', '')).strip().lower()
            if coper_name and coper_name in subject.lower():
                keywords = str(row.get('keywords', '')).split(';')
                for keyword in keywords:
                    if keyword.lower() in subject.lower():
                        keyword_path = row.get('keyword_path', '')
                        return keyword_path, False, True
                save_path = row.get('save_path', '')
                return save_path, False, True
        return None, None, False
    else:
        row = rows.iloc[0]
        keywords = str(row.get('keywords', '')).split(';')
        for keyword in keywords:
            if keyword.lower() in subject.lower():
                keyword_path = row.get('keyword_path', '')
                return keyword_path, False, True
        special_case_value = str(row.get('special_case', '')).strip().lower() == 'yes'
        save_path = row.get('save_path', '')
        return save_path, special_case_value, True

def update_excel_summary(date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails):
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Summary'
        sheet.append(['Date', 'Total Emails', 'Saved in Default', 'Saved in Actual Paths', 'Not Saved'])

    sheet = workbook.active
    sheet.append([date_str, total_emails, saved_default, saved_actual, not_saved])

    if 'Failed Emails' not in workbook.sheetnames:
        failed_sheet = workbook.create_sheet('Failed Emails')
        failed_sheet.append(['Date', 'Email Address', 'Subject'])
    else:
        failed_sheet = workbook['Failed Emails']

    for email in failed_emails:
        failed_sheet.append([date_str, email['email_address'], email['subject']])

    workbook.save(EXCEL_FILE_PATH)

def save_email(item, save_path, special_case):
    try:
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        
        valid_extensions = ('.xlsx', '.xls', '.csv', '.pdf', '.doc', '.docx')
        if special_case and item.Attachments.Count > 0:
            for attachment in item.Attachments:
                if attachment.FileName.lower().endswith(valid_extensions):
                    filename_base = sanitize_filename(os.path.splitext(attachment.FileName)[0])
                    break
            else:
                filename_base = sanitize_filename(item.Subject)
        else:
            filename_base = sanitize_filename(item.Subject)
        
        extension = ".msg"
        max_filename_length = 255 - len(save_path) - len(extension) - 1
        if len(filename_base) > max_filename_length:
            filename_base = filename_base[:max_filename_length]
        
        filename = f"{filename_base}{extension}"
        full_path = os.path.join(save_path, filename)
        
        # Check if a file with the same name already exists
        counter = 1
        while os.path.exists(full_path):
            filename = f"{filename_base}_{counter}{extension}"
            full_path = os.path.join(save_path, filename)
            counter += 1
        
        item.SaveAs(full_path, 3)
        return filename
    except pythoncom.com_error as com_err:
        error_message = f"COM Error saving email '{item.Subject}' to '{save_path}': {str(com_err)}"
        print(error_message)
        raise
    except Exception as e:
        error_message = f"General Error saving email '{item.Subject}' to '{save_path}': {str(e)}"
        print(error_message)
        raise

def process_email(item, sender_path_table, default_year, specific_date_str):
    logs = []
    failed_emails = []
    retries = 3
    processed = False

    # Attempt to extract sender email safely
    try:
        if hasattr(item, 'SenderEmailAddress') and item.SenderEmailAddress:
            sender_email = item.SenderEmailAddress.lower()
        elif hasattr(item, 'Sender') and item.Sender and hasattr(item.Sender, 'Address') and item.Sender.Address:
            sender_email = item.Sender.Address.lower()
        else:
            # If there's no sender info, skip
            logs.append(f"Skipped email '{item.Subject}' due to missing sender information.")
            return logs, failed_emails
    except Exception:
        logs.append(f"Skipped email '{item.Subject}' due to error fetching sender info.")
        return logs, failed_emails

    while retries > 0 and not processed:
        try:
            year, month = extract_date_from_text(item.Subject, default_year)
            if not year or not month:
                for attachment in item.Attachments:
                    year, month = extract_date_from_text(attachment.FileName, default_year)
                    if year and month:
                        break
            year = year or default_year

            base_path, special_case, is_csv_path = find_save_path(sender_email, item.Subject, sender_path_table)
            if base_path is None:
                base_path = DEFAULT_SAVE_PATH

            # If it's a default path email
            if not is_csv_path:
                save_path = os.path.join(base_path, specific_date_str)
            else:
                if special_case:
                    save_path = os.path.join(base_path, str(year))
                else:
                    if month:
                        save_path = os.path.join(base_path, str(year), month)
                    else:
                        save_path = os.path.join(base_path, str(year))

            print(f"Email from: {sender_email}")
            print(f"Subject: {item.Subject}")
            print(f"Special Case: {special_case}")
            print(f"Save Path: {save_path}")

            filename = save_email(item, save_path, special_case)
            logs.append(f"Saved: {filename} to {save_path}")
            processed = True
        except pythoncom.com_error as com_err:
            retries -= 1
            logs.append(f"COM Error handling email '{item.Subject}' from '{sender_email}' (Code: {com_err.args})")
            if retries == 0:
                logs.append(f"Failed to save the email '{item.Subject}' from '{sender_email}' after 3 retries")
                failed_emails.append({'email_address': sender_email, 'subject': item.Subject})
        except Exception as e:
            retries = 0
            logs.append(f"Error handling email '{item.Subject}' from '{sender_email}': {str(e)}")
            failed_emails.append({'email_address': sender_email, 'subject': item.Subject})

    return logs, failed_emails

def save_emails_from_senders_on_date(email_address, specific_date_str, sender_path_table, default_year):
    logs = []
    pythoncom.CoInitialize()
    specific_date = datetime.datetime.strptime(specific_date_str, '%Y-%m-%d').date()
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

    # Recursive search helper
    def find_folder_by_name(parent_folder, target_name):
        for f in parent_folder.Folders:
            if f.Name.lower() == target_name.lower():
                return f
            sub_result = find_folder_by_name(f, target_name)
            if sub_result is not None:
                return sub_result
        return None

    inbox = None
    malai_folder = None

    # Locate the store
    for store in outlook.Stores:
        if store.DisplayName.lower() == email_address.lower() or store.ExchangeStoreType == 3:
            try:
                root_folder = store.GetRootFolder()
                inbox = next((folder for folder in root_folder.Folders if folder.Name.lower() == "inbox"), None)

                if inbox:
                    # Find 'Malai' folder recursively under Inbox
                    malai_folder = find_folder_by_name(inbox, "malai")

                break
            except AttributeError as e:
                logs.append(f"Error accessing folders: {str(e)}")
                continue

    if not inbox:
        logs.append(f"No Inbox found for the account with email address: {email_address}")
        pythoncom.CoUninitialize()
        with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
            f.writelines("\n".join(logs))
        return
    else:
        logs.append("Inbox found successfully.")

    if not malai_folder:
        logs.append("No 'Malai' folder found as a subfolder of Inbox.")
    else:
        logs.append("Malai folder found successfully.")

    def get_items_for_folder(folder, date):
        filtered_items = []
        if folder:
            items = folder.Items
            items.Sort("[ReceivedTime]", True)
            items = items.Restrict(
                f"[ReceivedTime] >= '{date.strftime('%m/%d/%Y')} 00:00 AM' AND [ReceivedTime] <= '{date.strftime('%m/%d/%Y')} 11:59 PM'"
            )
            for item in items:
                filtered_items.append(item)
        return filtered_items

    inbox_items = get_items_for_folder(inbox, specific_date) if inbox else []
    malai_items = get_items_for_folder(malai_folder, specific_date) if malai_folder else []

    all_items = inbox_items + malai_items
    logs.append(f"Total emails found: {len(all_items)} (Inbox: {len(inbox_items)}, Malai: {len(malai_items)})")

    total_emails = len(all_items)
    saved_default, saved_actual, not_saved = 0, 0, 0
    failed_emails = []

    for item in all_items:
        email_logs, email_failed_emails = process_email(item, sender_path_table, default_year, specific_date_str)
        logs.extend(email_logs)
        failed_emails.extend(email_failed_emails)
        if any(DEFAULT_SAVE_PATH in log for log in email_logs):
            saved_default += 1
        else:
            saved_actual += 1

    pythoncom.CoUninitialize()
    with open(LOG_FILE_PATH, 'w', encoding='utf-8') as f:
        f.writelines("\n".join(logs))

    update_excel_summary(specific_date_str, total_emails, saved_default, saved_actual, not_saved, failed_emails)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        date_str = request.form['date']
        default_year = request.form['default_year']
        file = request.files['file']
        if file and date_str and default_year:
            try:
                datetime.datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                flash("Invalid date format. Please enter the date in YYYY-MM-DD format.", 'error')
                return redirect(url_for('index'))

            if not (default_year.isdigit() and len(default_year) == 4):
                flash("Invalid year format. Please enter the year in YYYY format.", 'error')
                return redirect(url_for('index'))

            filepath = os.path.join('uploads', sanitize_filename(file.filename))
            file.save(filepath)

            try:
                sender_path_table = pd.read_csv(filepath, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    sender_path_table = pd.read_csv(filepath, encoding='latin1')
                except Exception as e:
                    flash("Error reading the CSV file. Please ensure it's properly formatted.", 'error')
                    return redirect(url_for('index'))

            # Standardize column names to lowercase
            sender_path_table.columns = sender_path_table.columns.str.lower()

            account_email_address = "hf_data@bofa.com"
            socketio.start_background_task(save_emails_from_senders_on_date, account_email_address, date_str, sender_path_table, default_year)
            return redirect(url_for('results'))

    return render_template('index.html')

@app.route('/results')
def results():
    logs = []
    if os.path.exists(LOG_FILE_PATH):
        with open(LOG_FILE_PATH, 'r', encoding='utf-8') as f:
            logs = f.readlines()

    return render_template('results.html', logs=logs)

if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    os.makedirs(DEFAULT_SAVE_PATH, exist_ok=True)
    socketio.run(app, debug=True, use_reloader=False)
