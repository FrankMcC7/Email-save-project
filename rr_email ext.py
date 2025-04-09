It sounds like the issue is that the script is only detecting two accounts, even though the user has multiple inboxes in their Outlook. There could be a distinction between "accounts" and "inboxes" in Outlook that we need to address.

In Outlook, there are several ways inboxes can be organized:
1. Multiple email accounts (like separate Gmail, Outlook, Yahoo accounts)
2. Multiple folders within a single account
3. Shared mailboxes
4. Additional mailboxes added through Exchange

Let me rework the script to ensure it can capture all types of inboxes in Outlook, not just the primary accounts. I'll modify the approach to:

1. Get all available mailboxes/folders in Outlook regardless of how they're associated with accounts
2. Present a complete list to the user
3. Allow selection of multiple mailboxes/folders from this comprehensive list
4. Process each selected mailbox/folder

This should ensure we capture all inboxes that the user has access to in their Outlook application.

# Comprehensive Outlook Email Extractor

I've reworked the script to handle multiple inboxes in Outlook, including shared mailboxes and additional mailboxes that might not be appearing as separate accounts. This version should detect all mailboxes you have access to.

```python
import win32com.client
import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_sender_emails_from_outlook():
    """
    Connects to Outlook and extracts sender email addresses with subjects from
    multiple inboxes, including shared mailboxes and additional mailboxes.
    """
    try:
        # Connect to Outlook
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        
        # Get all available mailboxes/folders
        all_folders = []
        
        # First add default mailbox folders
        root_folder = namespace.Folders
        
        # Function to recursively get all mail folders
        def get_mail_folders(parent_folder, path=""):
            folders_list = []
            try:
                # Go through each folder in the parent
                for folder in parent_folder:
                    current_path = f"{path}/{folder.Name}" if path else folder.Name
                    
                    # Check if this is a mail folder we're interested in
                    try:
                        if folder.DefaultItemType == 0:  # 0 = Mail items
                            folders_list.append((current_path, folder))
                        
                        # Check if there are subfolders
                        if folder.Folders.Count > 0:
                            subfolders = get_mail_folders(folder.Folders, current_path)
                            folders_list.extend(subfolders)
                    except:
                        # Skip folders that throw errors
                        pass
            except:
                pass
            
            return folders_list
        
        # Get all mail folders
        all_folders = get_mail_folders(root_folder)
        
        if not all_folders:
            print("No mail folders found. Please ensure Outlook is properly configured.")
            return
        
        # Display all available folders
        print("\nAvailable mailboxes and folders:")
        print("-" * 70)
        for i, (folder_path, _) in enumerate(all_folders, 1):
            print(f"{i}. {folder_path}")
        
        # Let user select folders
        print("\nSelect folder numbers (comma-separated, e.g., 1,2,3 or 'all' for all folders):")
        folder_input = input("Folder numbers: ").strip()
        
        selected_folders = []
        if folder_input.lower() == 'all':
            selected_folders = all_folders
        else:
            folder_indices = [int(idx.strip()) - 1 for idx in folder_input.split(',')]
            for idx in folder_indices:
                if 0 <= idx < len(all_folders):
                    selected_folders.append(all_folders[idx])
                else:
                    print(f"Warning: Folder number {idx+1} is invalid and will be skipped.")
        
        if not selected_folders:
            print("No valid folders selected.")
            return
        
        print(f"\nSelected {len(selected_folders)} folder(s):")
        for folder_path, _ in selected_folders:
            print(f"- {folder_path}")
        
        # Calculate date 30 days ago
        thirty_days_ago = datetime.datetime.now() - datetime.timedelta(days=30)
        
        # Restriction for emails from the last 30 days
        restriction = "[ReceivedTime] >= '" + thirty_days_ago.strftime('%m/%d/%Y %H:%M %p') + "'"
        
        # List to store all email data
        email_data = []
        total_emails_processed = 0
        
        # Process each selected folder
        for folder_path, folder in selected_folders:
            try:
                print(f"\nProcessing folder: {folder_path}")
                
                # Get emails from the last 30 days
                emails = folder.Items.Restrict(restriction)
                emails.Sort("[ReceivedTime]", True)  # Sort by received time in descending order
                
                folder_email_count = 0
                
                # Extract the account name from the folder path
                # The first part of the path is typically the account name
                account_name = folder_path.split('/')[0]
                
                # Process each email
                for email in emails:
                    try:
                        sender_address = email.SenderEmailAddress
                        if sender_address and "@" in sender_address:
                            # Extract necessary information
                            received_time = email.ReceivedTime
                            subject = email.Subject
                            
                            # Add to our data list
                            email_data.append({
                                'account': account_name,
                                'folder': folder_path,
                                'sender': sender_address,
                                'subject': subject,
                                'received': received_time
                            })
                            
                            folder_email_count += 1
                            total_emails_processed += 1
                            
                            # Show progress every 25 emails
                            if folder_email_count % 25 == 0:
                                print(f"Processed {folder_email_count} emails in current folder...")
                    except Exception as e:
                        print(f"Error processing an email: {str(e)}")
                
                print(f"Processed {folder_email_count} emails in {folder_path}")
                
            except Exception as e:
                print(f"Error processing folder {folder_path}: {str(e)}")
        
        if not email_data:
            print("\nNo emails found in the selected folders from the last 30 days.")
            return
        
        # Ask for output file name
        default_filename = f"outlook_emails_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_file = input(f"\nEnter output Excel file name (default: {default_filename}): ") or default_filename
        
        # Ensure file has .xlsx extension
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'
        
        # Create Excel file
        create_excel_file(email_data, output_file)
        
        print(f"\nExtracted data from {total_emails_processed} emails across {len(selected_folders)} folder(s)")
        print(f"Results saved to {os.path.abspath(output_file)}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

def create_excel_file(email_data, output_file):
    """
    Creates an Excel file with the extracted email data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Senders"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Create headers
    headers = ["Account", "Folder", "Sender Email", "Subject", "Received Date"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    for row_num, data in enumerate(email_data, 2):
        ws.cell(row=row_num, column=1).value = data['account']
        ws.cell(row=row_num, column=2).value = data['folder']
        ws.cell(row=row_num, column=3).value = data['sender']
        ws.cell(row=row_num, column=4).value = data['subject']
        ws.cell(row=row_num, column=5).value = data['received'].strftime('%Y-%m-%d %H:%M')
        
        # Apply borders to all cells
        for col_num in range(1, 6):
            ws.cell(row=row_num, column=col_num).border = thin_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = min(adjusted_width, 50)  # Cap width at 50
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Create a summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    
    # Summary headers
    summary_headers = ["Account", "Folder", "Email Count"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Create summary data
    summary_data = {}
    for data in email_data:
        key = (data['account'], data['folder'])
        if key in summary_data:
            summary_data[key] += 1
        else:
            summary_data[key] = 1
    
    # Add summary data
    for row_num, ((account, folder), count) in enumerate(summary_data.items(), 2):
        ws_summary.cell(row=row_num, column=1).value = account
        ws_summary.cell(row=row_num, column=2).value = folder
        ws_summary.cell(row=row_num, column=3).value = count
        
        # Apply borders
        for col_num in range(1, 4):
            ws_summary.cell(row=row_num, column=col_num).border = thin_border
    
    # Auto-adjust column widths for summary
    for col in ws_summary.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = max_length + 2
        ws_summary.column_dimensions[column].width = min(adjusted_width, 50)
    
    # Freeze the header row in summary
    ws_summary.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(output_file)

if __name__ == "__main__":
    print("=" * 70)
    print("Comprehensive Outlook Email Extractor".center(70))
    print("=" * 70)
    print("\nThis script will find all mailboxes and folders in your Outlook")
    print("and extract sender emails from the last 30 days.")
    
    extract_sender_emails_from_outlook()
    
    print("\nDone!")
```

## Key Improvements:

1. **Comprehensive Folder Discovery**:
   - The script now recursively scans all folders in Outlook
   - Detects all mailboxes including shared mailboxes and additional inboxes
   - Captures the full folder hierarchy

2. **Simplified Selection Process**:
   - Shows a complete list of all available mailboxes/folders
   - Allows selecting any combination of folders from the list

3. **Accurate Folder Path Tracking**:
   - Maintains the full path of each folder for better identification
   - Preserves the mailbox/account information in the output

## How to Use:

1. Install the required packages:
   ```
   pip install pywin32 openpyxl
   ```

2. Make sure Outlook is open and running

3. Run the script and follow the prompts to:
   - Select from the comprehensive list of all available mailboxes and folders
   - Specify the output Excel file name

## Requirements:

- Windows operating system
- Outlook installed and running
- Python 3.x
- pywin32 library
- openpyxl library

This reworked version should detect all inboxes regardless of how they're configured in your Outlook, giving you access to shared mailboxes, additional mailboxes, and any other inbox folders that might not be appearing as separate accounts.